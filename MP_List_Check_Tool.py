#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：MP TOOL PROJECT
@File    ：MP List Check Tool.py
@Author  ：zxh
@Date    ：2025/3/19 16:20
'''
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import os
from openpyxl import Workbook
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox, scrolledtext
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import seaborn as sns
import matplotlib.animation as animation
from read_excel import *
from check import *
from analyze import *
from analyzeKS import *
from analyze_huanan_project import *

pd.options.mode.chained_assignment = None  # 禁用警告
# 解决中文显示问题
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

multiple = 1.5


# class EditableCombobox(ttk.Combobox):
#     def __init__(self, master=None, category=None, **kwargs):
#         values = kwargs.pop('values', [])
#         super().__init__(master, **kwargs)
#         self.category = category
#         self._original_values = values.copy()  # 初始值
#         self._excel_values = []  # 从Excel读取的值
#         self['values'] = values
#         self._edit_mode = False
#         self._last_valid_value = ""
#
#         self['values'] = self._original_values
#         self._edit_mode = False
#         self._last_valid_value = ""
#
#         # 绑定事件
#         self.bind('<KeyRelease>', self._on_key_release)
#         self.bind('<FocusOut>', self._on_focus_out)
#         self.bind('<Button-3>', self._on_right_click)
#         self.bind('<Escape>', self._cancel_edit)
#
#         self._update_state()
#         self._ensure_excel_file()
#         self._load_values_from_excel()
#
#     def _ensure_excel_file(self):
#         """确保Excel文件存在"""
#         if not os.path.exists('compare_information.xlsx'):
#             df = pd.DataFrame(columns=['IC', 'Module', 'Glass', 'Year', 'Grade', 'Type', 'Principal','Publisher'])
#             df.to_excel('compare_information.xlsx', index=False)
#
#     def _load_values_from_excel(self):
#         """从Excel文件加载已有选项"""
#         try:
#             excel_path = 'compare_information.xlsx'
#             if not os.path.exists(excel_path):
#                 return
#
#             df = pd.read_excel(excel_path, sheet_name='Sheet2')
#             if self.category and not df.empty and self.category in df.columns:
#                 # 保持Excel中的原始顺序，只去除空值
#                 self._excel_values = [v for v in df[self.category].tolist() if pd.notna(v)]
#
#                 # 合并初始值和Excel值，去除重复项但保持顺序
#                 combined = []
#                 seen = set()
#                 for v in self._original_values:
#                     if v not in seen:
#                         seen.add(v)
#                         combined.append(v)
#
#                 for v in self._excel_values:
#                     if v not in seen:
#                         seen.add(v)
#                         combined.append(v)
#
#                 self._original_values = combined
#                 self['values'] = combined
#
#         except Exception as e:
#             print(f"加载Excel选项失败: {e}")
#
#     def _update_state(self):
#         """更新控件状态"""
#         if self._edit_mode:
#             self.config(state="normal")
#             self.config(background="#FFFACD")  # 浅黄色背景表示编辑模式
#             self.config(foreground="black")
#             # 显示提示工具提示
#             self.tooltip = tk.Toplevel(self)
#             self.tooltip.wm_overrideredirect(True)
#             self.tooltip_label = tk.Label(self.tooltip, text="右键保存", bg="yellow", fg="black")
#             self.tooltip_label.pack()
#             # 定位提示框
#             x = self.winfo_rootx() + self.winfo_width()
#             y = self.winfo_rooty()
#             self.tooltip.wm_geometry(f"+{x}+{y}")
#         else:
#             self.config(state="readonly")
#             self.config(background="SystemButtonFace")
#             self.config(foreground="black")
#             # 移除工具提示
#             if hasattr(self, 'tooltip'):
#                 self.tooltip.destroy()
#
#     def _on_right_click(self, event=None):
#         """右键单击事件处理"""
#         if self._edit_mode:
#             # 如果当前是编辑模式，保存并退出
#             current_text = self.get()
#             if current_text and current_text not in self._original_values and self.category:
#                 # 添加到当前选项列表
#                 self._original_values.append(current_text)
#                 self['values'] = self._original_values
#                 self._save_to_excel(current_text)
#             elif current_text in self._original_values:
#                 # 如果是已有值，确保选中它
#                 self.set(current_text)
#             self._edit_mode = False
#         else:
#             # 如果当前是只读模式，进入编辑模式
#             self._edit_mode = True
#             self._last_valid_value = self.get()  # 记录当前值
#             self.icursor(tk.END)
#             self.selection_range(0, tk.END)
#             self.focus_set()
#
#         self._update_state()
#
#     def _on_key_release(self, event):
#         """按键释放事件处理 - 自动匹配已有选项"""
#         if not self._edit_mode:
#             return
#
#         current_text = self.get()
#         if current_text:
#             matches = [v for v in self._original_values if str(v).lower().startswith(current_text.lower())]
#             self['values'] = matches if matches else self._original_values
#         else:
#             self['values'] = self._original_values
#
#     def _on_focus_out(self, event):
#         """失去焦点事件处理"""
#         if self._edit_mode:
#             current_text = self.get()
#             if current_text and current_text not in self._original_values and self.category:
#                 # 添加到当前选项列表
#                 self._original_values.append(current_text)
#                 self['values'] = self._original_values
#                 self._save_to_excel(current_text)
#             elif current_text in self._original_values:
#                 # 如果是已有值，确保选中它
#                 self.set(current_text)
#             self._edit_mode = False
#             self._update_state()
#
#     def _cancel_edit(self, event=None):
#         """取消编辑"""
#         if self._edit_mode:
#             self._edit_mode = False
#             self.set(self._last_valid_value)  # 恢复上次有效值
#             self._update_state()
#
#     def _save_to_excel(self, new_value):
#         """保存新选项到Excel文件"""
#         try:
#             excel_path = 'compare_information.xlsx'
#
#             try:
#                 book = load_workbook(excel_path)
#                 if 'Sheet2' not in book.sheetnames:
#                     book.create_sheet('Sheet2')
#                     book.save(excel_path)
#             except:
#                 book = Workbook()
#                 if 'Sheet' in book.sheetnames:
#                     book['Sheet'].title = 'Sheet2'
#                 else:
#                     book.create_sheet('Sheet2')
#                 book.save(excel_path)
#
#             # 读取或创建DataFrame
#             try:
#                 df = pd.read_excel(excel_path, sheet_name='Sheet2')
#             except:
#                 df = pd.DataFrame()
#
#             if self.category not in df.columns:
#                 df[self.category] = pd.NA
#             # 查找该列的第一个空白行
#             empty_row_index = None
#             for idx, value in enumerate(df[self.category]):
#                 if pd.isna(value) or value == '':
#                     empty_row_index = idx
#                     break
#
#             # 如果有空白行，则在该行写入新值
#             if empty_row_index is not None:
#                 df.at[empty_row_index, self.category] = new_value
#             else:
#                 # 如果没有空白行，则在末尾添加新行
#                 new_row = {col: pd.NA for col in df.columns}
#                 new_row[self.category] = new_value
#                 df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
#             # 保存回Excel
#             with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
#                 df.to_excel(writer, sheet_name='Sheet2', index=False)
#
#         except Exception as e:
#             print(f"保存选项到Excel失败: {e}")

class EditableCombobox(ttk.Combobox):
    def __init__(self, master=None, category=None, sheet_name='Sheet2', **kwargs):
        values = kwargs.pop('values', [])
        super().__init__(master, **kwargs)
        self.category = category
        self.sheet_name = sheet_name
        self._original_values = values.copy()  # 初始值
        self._excel_values = []  # 从Excel读取的值
        self['values'] = values
        self._edit_mode = False
        self._last_valid_value = ""

        self['values'] = self._original_values
        self._edit_mode = False
        self._last_valid_value = ""

        # 绑定事件
        self.bind('<KeyRelease>', self._on_key_release)
        self.bind('<FocusOut>', self._on_focus_out)
        self.bind('<Button-3>', self._on_right_click)
        self.bind('<Escape>', self._cancel_edit)

        self._update_state()
        self._ensure_excel_file()
        self._load_values_from_excel()

    def _ensure_excel_file(self):
        """确保Excel文件存在并创建相应sheet的列"""
        excel_path = 'compare_information.xlsx'

        # 定义不同sheet的列
        sheet_columns = {
            'Sheet2': ['IC', 'Module', 'Glass', 'Year', 'Grade', 'Type', 'Principal', 'Publisher'],
            'Sheet3': ['IC_Type', 'Factory', 'Month', 'Glass', 'TP_Mode', 'Palm', 'Status', 'Category', 'Sales', 'PM',
                       'Year']
        }

        if not os.path.exists(excel_path):
            # 创建新Excel文件
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for sheet, columns in sheet_columns.items():
                    df = pd.DataFrame(columns=columns)
                    df.to_excel(writer, sheet_name=sheet, index=False)
        else:
            # 检查并添加缺失的sheet
            try:
                # 加载工作簿
                book = load_workbook(excel_path)

                # 检查并创建缺失的sheet
                for sheet, columns in sheet_columns.items():
                    if sheet not in book.sheetnames:
                        # 创建新的sheet
                        new_sheet = book.create_sheet(sheet)
                        # 写入列名
                        for col_idx, col_name in enumerate(columns, 1):
                            new_sheet.cell(row=1, column=col_idx, value=col_name)
                    else:
                        # 检查并添加缺失的列
                        sheet_obj = book[sheet]
                        # 获取现有列名（第一行）
                        existing_cols = []
                        max_col = sheet_obj.max_column
                        for col in range(1, max_col + 1):
                            cell_value = sheet_obj.cell(row=1, column=col).value
                            existing_cols.append(cell_value if cell_value else f'Column_{col}')

                        # 添加缺失的列
                        for col in columns:
                            if col not in existing_cols:
                                max_col += 1
                                sheet_obj.cell(row=1, column=max_col, value=col)

                # 保存工作簿
                book.save(excel_path)

            except Exception as e:
                print(f"检查Excel文件结构失败: {e}")
                # 如果openpyxl操作失败，尝试用pandas创建新文件
                try:
                    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                        for sheet, columns in sheet_columns.items():
                            # 尝试读取现有sheet，如果不存在则创建新的
                            try:
                                df = pd.read_excel(excel_path, sheet_name=sheet)
                            except:
                                df = pd.DataFrame()

                            # 确保所有需要的列都存在
                            for col in columns:
                                if col not in df.columns:
                                    df[col] = pd.NA

                            df.to_excel(writer, sheet_name=sheet, index=False)
                except Exception as e2:
                    print(f"使用pandas重新创建文件也失败: {e2}")

    def _load_values_from_excel(self):
        """从Excel文件指定sheet加载已有选项"""
        try:
            excel_path = 'compare_information.xlsx'
            if not os.path.exists(excel_path):
                return

            # 读取指定sheet的数据
            df = pd.read_excel(excel_path, sheet_name=self.sheet_name)
            if self.category and not df.empty and self.category in df.columns:
                # 保持Excel中的原始顺序，只去除空值
                self._excel_values = [str(v) for v in df[self.category].tolist() if
                                      pd.notna(v) and str(v).strip() != '']

                # 合并初始值和Excel值，去除重复项但保持顺序
                combined = []
                seen = set()
                for v in self._original_values:
                    v_str = str(v)
                    if v_str not in seen:
                        seen.add(v_str)
                        combined.append(v_str)

                for v in self._excel_values:
                    if v not in seen:
                        seen.add(v)
                        combined.append(v)

                self._original_values = combined
                self['values'] = combined

        except Exception as e:
            print(f"从Excel加载选项失败 (Sheet: {self.sheet_name}): {e}")

    def _update_state(self):
        """更新控件状态"""
        if self._edit_mode:
            self.config(state="normal")
            self.config(background="#FFFACD")  # 浅黄色背景表示编辑模式
            self.config(foreground="black")
            # 显示提示工具提示
            self.tooltip = tk.Toplevel(self)
            self.tooltip.wm_overrideredirect(True)
            self.tooltip_label = tk.Label(self.tooltip, text="右键保存", bg="yellow", fg="black")
            self.tooltip_label.pack()
            # 定位提示框
            x = self.winfo_rootx() + self.winfo_width()
            y = self.winfo_rooty()
            self.tooltip.wm_geometry(f"+{x}+{y}")
        else:
            self.config(state="readonly")
            self.config(background="SystemButtonFace")
            self.config(foreground="black")
            # 移除工具提示
            if hasattr(self, 'tooltip'):
                self.tooltip.destroy()

    def _on_right_click(self, event=None):
        """右键单击事件处理"""
        if self._edit_mode:
            # 如果当前是编辑模式，保存并退出
            current_text = self.get().strip()
            if current_text and current_text not in self._original_values and self.category:
                # 添加到当前选项列表
                self._original_values.append(current_text)
                self['values'] = self._original_values
                self._save_to_excel(current_text)
            elif current_text in self._original_values:
                # 如果是已有值，确保选中它
                self.set(current_text)
            self._edit_mode = False
        else:
            # 如果当前是只读模式，进入编辑模式
            self._edit_mode = True
            self._last_valid_value = self.get()  # 记录当前值
            self.icursor(tk.END)
            self.selection_range(0, tk.END)
            self.focus_set()

        self._update_state()

    def _on_key_release(self, event):
        """按键释放事件处理 - 自动匹配已有选项"""
        if not self._edit_mode:
            return

        current_text = self.get()
        if current_text:
            matches = [v for v in self._original_values if str(v).lower().startswith(current_text.lower())]
            self['values'] = matches if matches else self._original_values
        else:
            self['values'] = self._original_values

    def _on_focus_out(self, event):
        """失去焦点事件处理"""
        if self._edit_mode:
            current_text = self.get().strip()
            if current_text and current_text not in self._original_values and self.category:
                # 添加到当前选项列表
                self._original_values.append(current_text)
                self['values'] = self._original_values
                self._save_to_excel(current_text)
            elif current_text in self._original_values:
                # 如果是已有值，确保选中它
                self.set(current_text)
            self._edit_mode = False
            self._update_state()

    def _cancel_edit(self, event=None):
        """取消编辑"""
        if self._edit_mode:
            self._edit_mode = False
            self.set(self._last_valid_value)  # 恢复上次有效值
            self._update_state()

    def _save_to_excel(self, new_value):
        """保存新选项到Excel文件指定sheet"""
        try:
            excel_path = 'compare_information.xlsx'

            # 确保文件存在并创建需要的sheet
            self._ensure_excel_file()

            # 读取指定sheet的数据
            try:
                df = pd.read_excel(excel_path, sheet_name=self.sheet_name)
            except:
                # 如果读取失败，创建空的DataFrame
                df = pd.DataFrame()

            # 确保列存在
            if self.category not in df.columns:
                df[self.category] = pd.NA

            # 查找该列的第一个空白行
            empty_row_index = None
            for idx, value in enumerate(df[self.category]):
                if pd.isna(value) or str(value).strip() == '':
                    empty_row_index = idx
                    break

            # 如果有空白行，则在该行写入新值
            if empty_row_index is not None:
                df.at[empty_row_index, self.category] = new_value
            else:
                # 如果没有空白行，则在末尾添加新行
                new_row = {col: pd.NA for col in df.columns}
                new_row[self.category] = new_value
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)

            # 保存回Excel的指定sheet
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=self.sheet_name, index=False)

            print(f"成功保存到 {self.sheet_name} 的 {self.category} 列: {new_value}")

        except Exception as e:
            print(f"保存选项到Excel失败 (Sheet: {self.sheet_name}): {e}")

    def set_sheet(self, sheet_name):
        """切换sheet并重新加载数据"""
        self.sheet_name = sheet_name
        self._original_values = []
        self._excel_values = []
        self._load_values_from_excel()
        self['values'] = self._original_values
        self.set('')  # 清空当前值



class CheckFrame(tk.Frame):
    """
    Check界面
    """

    def __init__(self, master=None):
        """
        self值初始化
        :param master: NONE
        """
        super().__init__(master)
        self.master = master
        self.file_path = None
        self.df = None
        self.valid_glass_models = None
        self.valid_TOOL_version = None
        self.valid_IC = None
        self.checks_visible = False  # 用于跟踪功能按钮的显示状态
        self.create_widgets()
        self.sequence_error = None
        self.position_error = None

    def create_widgets(self):
        """
        绘制UI界面
        :return:
        """
        # 文件路径显示区域
        self.label_file_path = tk.Label(self, text="文 件 路 径:")
        self.label_file_path.place(x=10, y=10)

        self.text_file_path = tk.Entry(self, state="readonly")
        self.text_file_path.place(x=90, y=10, width=320 * multiple + 20, height=25)

        # 文件选择按钮
        self.btn_open = tk.Button(self, text="...", width=3, bg='Gainsboro', command=self.open_file)
        self.btn_open.place(x=420 * multiple, y=10)

        # 重新读取文件按钮
        self.reload_var = tk.BooleanVar(value=True)  # 默认勾选
        self.chk_reload = tk.Checkbutton(self, text="reload excel", width=15, variable=self.reload_var)
        self.chk_reload.place(x=460 * multiple, y=10)

        # self.btn_reload_excel = tk.Button(self, text="重新读取文件", height=1, width=15, command=self.reload_excel,
        #                                   state=tk.DISABLED)
        # self.btn_reload_excel.place(x=90, y=50)

        # 一键检查按钮
        self.btn_check_all = tk.Button(self, text="一键检查", height=1, width=15, command=self.run_check_all,
                                       state=tk.DISABLED)
        self.btn_check_all.place(x=80, y=50)

        # 分项检查按钮
        self.btn_show_checks = tk.Button(self, text="分项检查", height=1, width=15, command=self.show_checks,
                                         state=tk.DISABLED)
        self.btn_show_checks.place(x=250 * multiple, y=50)

        # 序号校正按钮
        self.btn_sequence_revise = tk.Button(self, text="序号校正", height=1, width=15,
                                             command=lambda: self.fix_sequence_errors(self.sequence_error),
                                             state=tk.DISABLED)
        self.btn_sequence_revise.place(x=420 * multiple, y=50)

        # 检查按钮（初始状态为隐藏和禁用）
        self.btn_sequence = tk.Button(self, text="序列检查", height=1, width=15,
                                      command=lambda: self.run_check("序列检查", lambda: check_sequence(self.df)))
        self.btn_grade = tk.Button(self, text="等级检查", height=1, width=15,
                                   command=lambda: self.run_check("等级检查", lambda: check_grade(self.df.iloc[:, 1])))
        self.btn_ic_model = tk.Button(self, text="IC型号检查", height=1, width=15, command=lambda: self.run_check("IC型号检查",
                                                                                                              lambda: check_ic_model(
                                                                                                                  self.df.iloc[
                                                                                                                  :, 6],
                                                                                                                  self.valid_IC)))
        self.btn_glass_model = tk.Button(self, text="玻璃型号检查", height=1, width=15,
                                         command=lambda: self.run_check("玻璃型号检查",
                                                                        lambda: check_glass_model(self.df.iloc[:, 7],
                                                                                                  self.valid_glass_models)))
        self.btn_interface = tk.Button(self, text="接口定义检查", height=1, width=15, command=lambda: self.run_check("接口定义检查",
                                                                                                               lambda: check_interface(
                                                                                                                   self.df.iloc[
                                                                                                                   :,
                                                                                                                   8])))
        self.btn_flash = tk.Button(self, text="flash检查", height=1, width=15, command=lambda: self.run_check("flash检查",
                                                                                                            lambda: check_flash(
                                                                                                                self.df.iloc[
                                                                                                                :, 9],
                                                                                                                self.df.iloc[
                                                                                                                :, 8])))
        self.btn_tool_version = tk.Button(self, text="TOOL版本检查", height=1, width=15,
                                          command=lambda: self.run_check("TOOL版本检查", lambda: check_tool_version(self.df,
                                                                                                                self.valid_TOOL_version)))
        self.btn_date_format = tk.Button(self, text="发布日期检查", height=1, width=15,
                                         command=lambda: self.run_check("发布日期检查", lambda: check_date_format(self.df)))

        self.btn_project_name = tk.Button(self, text="项目号是否重复检查", height=1, width=15,
                                          command=lambda: self.run_check("项目号是否重复检查",
                                                                         lambda: check_same_project(self.df)))
        self.btn_project_revise = tk.Button(self, text="项目号位置校正", height=1, width=15,
                                            command=lambda: self.move_same_projects_position(self.position_error))

        # 结果显示区域
        self.result_area = scrolledtext.ScrolledText(self)
        self.result_area.place(x=10 * multiple, y=90, width=590 * multiple, height=480 * multiple)
        self.result_area.config(state=tk.DISABLED)

        # 设置颜色标签
        self.result_area.tag_config("green", foreground="green")
        self.result_area.tag_config("red", foreground="red")

    def open_file(self):
        """
        打开文件夹并且写入文件路径，用于打开文件操作
        :return:
        """
        # 打开文件选择对话框
        self.file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])
        # 选择文件路径后，清一次屏
        self.result_area.config(state=tk.NORMAL)
        self.result_area.delete(1.0, tk.END)
        self.result_area.config(state=tk.DISABLED)
        if self.file_path:
            # 更新文件路径
            self.text_file_path.config(state="normal")
            self.text_file_path.delete(0, tk.END)
            self.text_file_path.insert(0, self.file_path)
            self.text_file_path.config(state="readonly")

            # 启用一键检查和分项检查按钮
            self.btn_check_all.config(state=tk.NORMAL)
            self.btn_show_checks.config(state=tk.NORMAL)

    def reload_excel(self):
        """
        加载文件，用于重新加载文件按键操作
        :return:
        """
        try:
            self.df = read_excel_with_merged_cells(self.file_path, sheet_name='MP Project List（internal）')
            if self.df is not None:
                # 读取对比文件
                valid_data_file = 'compare_information.xlsx'
                try:
                    valid_data = read_compare_excel(valid_data_file, sheet_names='Sheet1')
                    if valid_data is not None and not valid_data.empty:
                        self.valid_TOOL_version = valid_data.iloc[:, 1].dropna().tolist()
                        self.valid_glass_models = valid_data.iloc[:, 2].dropna().tolist()
                        self.valid_IC = valid_data.iloc[:, 0].dropna().tolist()
                    else:
                        self.valid_TOOL_version = []
                        self.valid_glass_models = []
                        self.valid_IC = []
                        messagebox.showwarning("警告", "对比文件为空或无效")
                except Exception as e:
                    self.valid_TOOL_version = []
                    self.valid_glass_models = []
                    self.valid_IC = []
                    messagebox.showerror("错误", f"读取对比文件失败: {e}")
            else:
                messagebox.showerror("错误", "读取Excel文件失败")
        except Exception as e:
            self.df = None
            messagebox.showerror("错误", f"读取Excel文件失败: {e}")

    def show_checks(self):
        """
        隐藏/显示分项检查功能按键
        :return:
        """
        if self.checks_visible:
            # 隐藏功能按钮
            self.result_area.place(x=10 * multiple, y=90, width=590 * multiple, height=480 * multiple)
            self.btn_sequence.place_forget()
            self.btn_grade.place_forget()
            self.btn_ic_model.place_forget()
            self.btn_glass_model.place_forget()
            self.btn_interface.place_forget()
            self.btn_flash.place_forget()
            self.btn_tool_version.place_forget()
            self.btn_date_format.place_forget()
            self.btn_project_name.place_forget()
            self.btn_project_revise.place_forget()
            self.checks_visible = False
        else:
            # 显示功能按钮
            self.result_area.place(x=10 * multiple, y=170, width=590 * multiple, height=400 * multiple)
            self.btn_sequence.place(x=10 * multiple, y=90)
            self.btn_grade.place(x=120 * multiple, y=90)
            self.btn_ic_model.place(x=230 * multiple, y=90)
            self.btn_glass_model.place(x=340 * multiple, y=90)
            self.btn_interface.place(x=450 * multiple, y=90)
            self.btn_flash.place(x=10 * multiple, y=130)
            self.btn_tool_version.place(x=120 * multiple, y=130)
            self.btn_date_format.place(x=230 * multiple, y=130)
            self.btn_project_name.place(x=340 * multiple, y=130)
            self.btn_project_revise.place(x=450 * multiple, y=130)
            self.btn_project_revise.config(state=tk.DISABLED)
            self.checks_visible = True

    def run_check_all(self):
        """
        一次性检查所有项，用于一键检查按键操作
        :return:
        """
        if self.reload_var.get():  # 如果勾选框被勾选，默认重新加载文件
            self.reload_excel()
        checks = [
            ("序列检查", check_sequence(self.df)),
            ("等级检查", check_grade(self.df.iloc[:, 1])),
            ("IC型号检查", check_ic_model(self.df.iloc[:, 6], self.valid_IC)),
            ("玻璃型号检查", check_glass_model(self.df.iloc[:, 7], self.valid_glass_models)),
            ("接口定义检查", check_interface(self.df.iloc[:, 8])),
            ("flash有无检查", check_flash(self.df.iloc[:, 9], self.df.iloc[:, 8])),
            ("Tool版本号检查", check_tool_version(self.df, self.valid_TOOL_version)),
            ("发布日期检查", check_date_format(self.df)),
            ("项目号是否重复检查", check_same_project(self.df))
        ]
        # 清空结果显示区域
        self.result_area.config(state=tk.NORMAL)
        self.result_area.delete(1.0, tk.END)

        # 显示检查结果
        for check_name, errors in checks:
            self.result_area.insert(tk.END, f"{check_name}:\n")
            if not errors:
                self.result_area.insert(tk.END, "  PASS\n", "green")
                if check_name == "序列检查":
                    self.btn_sequence_revise.config(state=tk.DISABLED)
                if check_name == '项目号是否重复检查':
                    self.btn_project_revise.config(state=tk.DISABLED)
            else:
                self.result_area.insert(tk.END, "FAIL:\n", "red")
                self.result_area.insert(tk.END, "错误位置:\n", "red")
                for error in errors:
                    if check_name == "项目号是否重复检查":
                        self.result_area.insert(tk.END, f"项目号‘{error[3]}’,重复位置: 行 {error[0]} 和行 {error[1]} \n", "red")
                    else:
                        self.result_area.insert(tk.END, f" 第 {error[0]} 行, 错误值: {error[2]}\n", "red")

                # 序列检查有问题则启用序号校正按钮
                if check_name == "序列检查":
                    self.btn_sequence_revise.config(state=tk.NORMAL)
                    self.sequence_error = check_sequence_for_fix(self.df)

                # 项目号检查有问题启用项目号位置校正
                if check_name == '项目号是否重复检查':
                    self.btn_project_revise.config(state=tk.NORMAL)
                    self.position_error = check_same_project(self.df)

            self.result_area.insert(tk.END, "\n")

        self.result_area.config(state=tk.DISABLED)

    def run_check(self, check_name, check_func):
        """
        用于分项检查按键操作
        :param check_name: check项目名
        :param check_func: 对应check函数
        :return:
        """
        if self.reload_var.get():  # 如果勾选框被勾选，默认重新加载文件
            self.reload_excel()
        errors = check_func()
        # 清空结果显示区域
        self.result_area.config(state=tk.NORMAL)
        self.result_area.delete(1.0, tk.END)

        # 显示检查结果
        self.result_area.insert(tk.END, f"{check_name}:\n")
        if not errors:
            self.result_area.insert(tk.END, "  PASS\n", "green")
            if check_name == "序列检查":
                self.btn_sequence_revise.config(state=tk.DISABLED)
            if check_name == '项目号是否重复检查':
                self.btn_project_revise.config(state=tk.DISABLED)
        else:
            self.result_area.insert(tk.END, "FAIL:\n", "red")
            self.result_area.insert(tk.END, "错误位置:\n", "red")
            for error in errors:
                if check_name == "项目号是否重复检查":
                    self.result_area.insert(tk.END, f"项目号‘{error[3]}’,重复位置: 行 {error[0]}和行 {error[1]} \n", "red")
                else:
                    self.result_area.insert(tk.END, f" 第 {error[0]} 行, 错误值: {error[2]}\n", "red")

            # 序列检查有问题则启用序号校正按钮
            if check_name == "序列检查":
                self.btn_sequence_revise.config(state=tk.NORMAL)
                self.sequence_error = check_sequence_for_fix(self.df)

            # 项目号检查有问题启用项目号位置校正
            if check_name == '项目号是否重复检查':
                self.btn_project_revise.config(state=tk.NORMAL)
                self.position_error = check_same_project(self.df)

        self.result_area.insert(tk.END, "\n")

        self.result_area.config(state=tk.DISABLED)

    def fix_sequence_errors(self, errors):
        """
        修正序列错误，并保存到原文件的合并单元格中
        :param errors: 错误列表，格式为 [(行号, 列名, 错误值), ...]
        :return:
        """
        try:
            fix_sequence(self.file_path, errors)
            messagebox.showinfo("success", "序列已校正并保存到modified_excel_file.xlsx！")
        except Exception as e:
            messagebox.showerror("fail!", f"校正序列失败{e}")

    def move_same_projects_position(self, errors):
        """
        :param errors:
        :return:
        """
        try:
            result_path, merge_ranges=cut_and_paste_excel(self.file_path, errors)
            file_path_moved_position = "cut_paste_result.xlsx"
            try:
                result_file = merge_and_center_cells(
                    workbook_path=file_path_moved_position,
                    sheet_name="MP Project List（internal）",
                    merge_ranges=merge_ranges,
                    output_path="excel_style_merged.xlsx"
                )
                print("合并完成！")
            except Exception as e:
                print(f"合并过程中出错: {e}")
                messagebox.showerror("fail!", f"项目单元格合并失败{e}")

            file_path_merged ="excel_style_merged.xlsx"
            try:
                self.df = read_excel_with_merged_cells(file_path_merged, sheet_name='MP Project List（internal）')
                self.sequence_error = check_sequence_for_fix(self.df)
                fix_sequence(file_path_merged, self.sequence_error)
                os.remove(file_path_moved_position)
                os.remove(file_path_merged)
            except Exception as e:
                messagebox.showerror("fail!", f"项目序号校正失败{e}")

            messagebox.showinfo("success", "项目位置已校正并保存到modified_excel_file.xlsx！！")
        except Exception as e:
            messagebox.showerror("fail!", f"校正项目信息位置失败{e}")


class AnalyseFrame(tk.Frame):
    """
    Analyse界面
    """

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master

        # 创建内部标签页容器
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill='both', expand=True)

        # 自定义更小的标签页样式
        style = ttk.Style()

        # 创建自定义标签样式
        style.configure('Small.TNotebook.Tab',
                        padding=(3, 1),
                        font=('Arial', 10, 'normal'),
                        width=8,
                        anchor='center')

        # 应用自定义样式
        self.notebook.configure(style='Small.TNotebook')

        # 创建MP分析标签页
        self.mp_frame = tk.Frame(self.notebook)
        self.notebook.add(self.mp_frame, text='MP分析', padding=0)

        # 创建客诉分析标签页
        self.ks_frame = tk.Frame(self.notebook)
        self.notebook.add(self.ks_frame, text='客诉分析', padding=0)

        #创建华南分析标签页
        self.HN_frame = tk.Frame(self.notebook)
        self.notebook.add(self.HN_frame, text='华南项目', padding=0)

        # 初始化两个子标签页
        self.init_mp_frame()
        self.init_ks_frame()
        self.init_HN_frame()

        # 将子分析类pack到各自的frame中
        self.mp_analysis.pack(fill='both', expand=True)
        self.ks_analysis.pack(fill='both', expand=True)
        self.HN_analysis.pack(fill='both', expand=True)

        # 绑定标签页切换事件
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

    def init_mp_frame(self):
        """初始化MP分析标签页"""
        self.mp_analysis = MPAnalysis(self.mp_frame)

    def init_ks_frame(self):
        """初始化客诉分析标签页"""
        self.ks_analysis = KSAnalysis(self.ks_frame)

    def init_HN_frame(self):
        """初始化客诉分析标签页"""
        self.HN_analysis = HNAnalysis(self.HN_frame)

    def on_tab_changed(self, event):
        """标签页切换"""
        selected_tab = self.notebook.select()
        tab_text = self.notebook.tab(selected_tab, "text")

        if tab_text.strip() == "MP分析":
            if hasattr(self.mp_analysis, 'canvas') and self.mp_analysis.canvas:
                self.mp_analysis.canvas.get_tk_widget().place(x=0, y=120)
        elif tab_text.strip() == "客诉分析":
            if hasattr(self.ks_analysis, 'canvas_KS') and self.ks_analysis.canvas_KS:
                self.ks_analysis.canvas_KS.get_tk_widget().place(x=0, y=120)
        elif tab_text.strip() == "华南项目":
            if hasattr(self.HN_analysis, 'canvas_HN') and self.HN_analysis.canvas_HN:
                self.HN_analysis.canvas_HN.get_tk_widget().place(x=0, y=150)


class MPAnalysis(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.create_widgets()
        self.file_path_analyse = None
        self.df1 = None
        self.canvas = None  # 用于保存画布对象
        self.ic_values = None
        self.module_values = None
        self.glass_values = None
        self.flash_values = None
        self.year_values = None
        # 用于模组厂分析分页
        self.current_page = 0
        self.total_pages = 1
        self.module_data = None
        self.module_title_conditions = None
        self.animation = None  # 用于存储动画对象
        self.fig = None  # 用于存储图形对象
        self.ax = None  # 用于存储坐标轴对象

    def create_widgets(self):
        """

        :return:
        """
        # 文件路径显示区域
        self.label1_file_path = tk.Label(self, text="文 件 路 径:")
        self.label1_file_path.place(x=10, y=10)

        self.text1_file_path = tk.Entry(self, state="readonly")
        self.text1_file_path.place(x=90, y=10, width=320 * multiple + 30, height=25)

        # 文件选择按钮
        self.btn_open_analyse = tk.Button(self, text="...", width=3, bg='Gainsboro', command=self.analyse_open_file)
        self.btn_open_analyse.place(x=410 * multiple, y=10)

        # 将数据分析及其下拉框框起来
        group_frame = tk.LabelFrame(self, bd=2, relief="groove", padx=5, pady=5)
        group_frame.place(x=450 * multiple, y=10, width=130* multiple, height=90)
        # 数据分析按键
        self.btn_analyse = tk.Button(self, text="项目统计", command=self.get_xl_to_analyse, state=tk.DISABLED)
        self.btn_analyse.place(x=465 * multiple, y=15, width=100* multiple)

        self.analyse_values = ["ALL", 'IC型号', '模组厂', '玻璃厂', 'Flash', '年度', '等级', '发布人']
        self.analyse_dropdown = ttk.Combobox(self, values=self.analyse_values, state="readonly")
        self.analyse_dropdown.current(0)
        self.analyse_dropdown.place(x=465 * multiple, y=60, width=100* multiple)

        self.analyse_dropdown.bind("<<ComboboxSelected>>", self.choose_status)

        # 下拉框区域
        self.label_ic = tk.Label(self, text="IC型号:", anchor="e")
        self.label_ic.place(x=10 * multiple, y=50, width=50)
        # self.ic_values = ["ALL", '7272', '7202H', '7202M', '7302']
        self.ic_values = ['ALL']
        # self.ic_dropdown = ttk.Combobox(self, values=self.ic_values, state="readonly")
        self.ic_dropdown = EditableCombobox(self, category='IC', values=self.ic_values)
        self.ic_dropdown.current(0)
        self.ic_dropdown.place(x=45 * multiple, y=50, width=90)

        self.label_module = tk.Label(self, text="模组厂:", anchor="e")
        self.label_module.place(x=120 * multiple, y=50, width=50)
        # self.module_values = ['ALL', '海菲', '信利', '易快来', '联创', '创维', '同兴达', '海盛捷', '壹星', '精卓', '三龙', '合力泰', '华显', '维立',
        #                       '亿华', '立德', '晶泰', '德智欣', '中光电', '沛宏', '欣欣光电', '众铭安', '天正达', '瑞恒光电', '清创高', '汉龙时代',
        #                       '晶胜通', '龙煜', '金宏光电', '威达光电', '惠科', '华视', '正金晶光电', '华映', '长信新显', '宏凯',
        #                       '泰启', '百业', '共赢', '德实', '京龙', '如新电子', '皓显', '大通显示', '高展', '康华显通', '煜鑫', '宏利超显', '菲触显视',
        #                       '轩达', '钜沣', '鹰芒技术', '德普特', '元格', '亿普拉斯', '万联', '重联', '日日佳', '中正威', '天山电子', '中正威', '天正达',
        #                       '凯晟', '联信康','美力凯']
        self.module_values = ['ALL']
        # self.module_dropdown = ttk.Combobox(self, values=self.module_values, state="readonly")
        self.module_dropdown = EditableCombobox(self, category='Module', values=self.module_values)
        self.module_dropdown.current(0)
        self.module_dropdown.place(x=155 * multiple, y=50, width=90)

        self.label_glass = tk.Label(self, text="玻璃厂:", anchor="e")
        self.label_glass.place(x=230 * multiple, y=50, width=50)
        # self.glass_values = ['ALL', 'CSOT', 'TM', 'TRULY', 'BOE', 'CTO', 'PANDA', 'SHARP', 'MDT', 'HKC', 'HSD', 'INX',
        #                      'IVO', 'CTC']
        self.glass_values = ['ALL']
        # self.glass_dropdown = ttk.Combobox(self, values=self.glass_values, state="readonly")
        self.glass_dropdown = EditableCombobox(self, category='Glass', values=self.glass_values)
        self.glass_dropdown.current(0)
        self.glass_dropdown.place(x=265 * multiple, y=50, width=90)

        self.label_flash = tk.Label(self, text="Flash:", anchor="e")
        self.label_flash.place(x=335 * multiple, y=50, width=50)
        self.flash_values = ['ALL', 'Y', 'N']
        self.flash_dropdown = ttk.Combobox(self, values=self.flash_values, state="readonly")
        self.flash_dropdown.current(0)
        self.flash_dropdown.place(x=370 * multiple, y=50, width=90)

        self.label_year = tk.Label(self, text="年度:", anchor="w")
        self.label_year.place(x=20* multiple, y=80, width=50)
        # self.year_values = ['ALL', '2022', '2023', '2024', '2025']
        self.year_values = ['ALL']
        # self.year_dropdown = ttk.Combobox(self, values=self.year_values, state="readonly")
        self.year_dropdown = EditableCombobox(self, category='Year', values=self.year_values)
        self.year_dropdown.current(0)
        self.year_dropdown.place(x=45 * multiple, y=80, width=90)

        self.label_grade = tk.Label(self, text="等级:", anchor="e")
        self.label_grade.place(x=120 * multiple, y=80, width=50)
        self.grade_values = ['ALL', 'NULL']
        # self.grade_dropdown = ttk.Combobox(self, values=self.grade_values, state="readonly")
        self.grade_dropdown = EditableCombobox(self, category='Grade', values=self.grade_values)
        self.grade_dropdown.current(0)
        self.grade_dropdown.place(x=155 * multiple, y=80, width=90)

        self.label_publisher = tk.Label(self, text="发布人:", anchor="e")
        self.label_publisher.place(x=230 * multiple, y=80, width=50)
        self.publisher_values = ['ALL']
        self.publisher_dropdown = EditableCombobox(self, category='Publisher', values=self.publisher_values)
        self.publisher_dropdown.current(0)
        self.publisher_dropdown.place(x=265 * multiple, y=80, width=90)

        # 模组分页设置
        # 初始化分页控件
        self.page_label = tk.Label(self)
        self.btn_prev = tk.Button(self, text="◀", width=2)
        self.btn_next = tk.Button(self, text="▶", width=2)

        # # 初始隐藏分页控件
        # self.page_label.place_forget()
        # self.btn_prev.place_forget()
        # self.btn_next.place_forget()

        # self.edit_tip = tk.Label(self, text="右键点击下拉框可添加新选项", fg="gray")
        # self.edit_tip.place(x=10, y=110)

    def analyse_open_file(self):
        """

        :return:
        """
        # 打开文件选择对话框
        self.file_path_analyse = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])

        if self.file_path_analyse:
            # 更新文件路径
            self.text1_file_path.config(state="normal")
            self.text1_file_path.delete(0, tk.END)
            self.text1_file_path.insert(0, self.file_path_analyse)
            self.text1_file_path.config(state="readonly")

            # 启用数据分析键
            self.btn_analyse.config(state=tk.NORMAL)

            # 重新选择文件则清除画布
            self.clear_plot()

    def analyse_reload_excel(self):
        """
        加载文件
        :return:
        """
        try:
            self.df1 = read_excel_for_analyse(self.file_path_analyse, sheet_name='MP Project List（internal）')
            # # 读取对比文件
            # valid_data_file = 'compare_information.xlsx'
            # try:
            #     valid_data = read_compare_excel(valid_data_file, sheet_names='Sheet2')
            #     if valid_data is not None and not valid_data.empty:
            #         TARGET_ICS = valid_data.iloc[:, 0].dropna().astype(str).tolist()
            #         FACTORIES = valid_data.iloc[:, 1].dropna().tolist()
            #         GLASS_TYPES = valid_data.iloc[:, 2].dropna().tolist()
            #         YEARS = valid_data.iloc[:, 3].dropna().tolist()
            #         GRADES = valid_data.iloc[:, 4].dropna().tolist()
            #         GRADES.append('NULL')
            #         KS_TYPES = valid_data.iloc[:, 5].dropna().tolist()
            #         PRINCIPAL = valid_data.iloc[:, 6].dropna().tolist()
            #     else:
            #         messagebox.showwarning("警告", "对比文件为空或无效")
            # except Exception as e:
            #     messagebox.showerror("错误", f"读取对比文件失败: {e}")
        except Exception as e:
            self.df1 = None
            messagebox.showerror("错误", f"读取Excel文件失败: {e}")

    def choose_status(self, event=None):
        choose = self.analyse_dropdown.get()
        if choose == 'ALL':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == 'IC型号':
            self.ic_dropdown.config(state=tk.DISABLED)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == '模组厂':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state=tk.DISABLED)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")


        elif choose == '玻璃厂':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state=tk.DISABLED)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == 'Flash':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state=tk.DISABLED)
            self.year_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == '年度':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state=tk.DISABLED)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == '等级':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state=tk.DISABLED)
            self.publisher_dropdown.config(state=tk.NORMAL)
            self.publisher_dropdown.config(state="readonly")

        elif choose == '发布人':
            self.ic_dropdown.config(state=tk.NORMAL)
            self.module_dropdown.config(state=tk.NORMAL)
            self.glass_dropdown.config(state=tk.NORMAL)
            self.flash_dropdown.config(state=tk.NORMAL)
            self.year_dropdown.config(state=tk.NORMAL)
            self.grade_dropdown.config(state=tk.NORMAL)
            self.ic_dropdown.config(state="readonly")
            self.module_dropdown.config(state="readonly")
            self.glass_dropdown.config(state="readonly")
            self.flash_dropdown.config(state="readonly")
            self.year_dropdown.config(state="readonly")
            self.grade_dropdown.config(state="readonly")
            self.publisher_dropdown.config(state=tk.DISABLED)

    def draw_histogram(self, choose, title, counts):
        """
        绘制直方图
        :param choose: X粥
        :param title: 筛选条件
        :param counts: 直方图数据
        :return: 直方图
        """
        # 绘制图表
        if int(counts.sum()) != 0:
            fig = plt.figure(figsize=(6*multiple-0.1, 4.6*multiple-0.05), dpi=100)
            f_plot = fig.add_subplot(111)  # 划分区域
            self.canvas = FigureCanvasTkAgg(fig, master=self)
            self.canvas.get_tk_widget().place(x=0, y=120)  # 放置位置
            f_plot.clear()  # 刷新

            bars = plt.bar(counts.index, counts.values, width=0.6)
            # 添加数据标签
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{int(height)}',
                         ha='center', va='bottom', fontsize=11, weight='bold')

            a_title = choose + "项目数量统计"
            if title:
                a_title += "\n(" + " | ".join(title) + ")"

            plt.title(a_title, fontsize=14, pad=10)
            plt.xlabel(choose, fontsize=12)
            plt.ylabel('MP项目数量', fontsize=12)
            if choose == '模组厂':
                plt.xticks(rotation=90, fontsize=10)
            else:
                plt.xticks(fontsize=10)
            # plt.yticks(fontsize=10)

            # 自动调整y轴范围，留出标签空间
            max_value = max(counts.values)
            plt.yticks(np.arange(0, max_value * 1.15 + 1, step=max(1, int(max_value / 10))), fontsize=10)
            plt.ylim(0, max_value * 1.15)
        else:
            messagebox.showinfo("提示", "该筛选条件下项目数量为0")

    def get_xl_to_analyse(self):
        """
        根据不同选择情况绘制直方图
        :return:
        """
        try:
            self.analyse_reload_excel()  # 每次点击数据分析可重新加载文件
            choose = self.analyse_dropdown.get()
            self.clear_plot()  # 每次选择先清除上一次直方图

            # 重置分页状态
            self.current_page = 0
            self.total_pages = 1
            # 其他分析类型不需要分页
            self.update_page_buttons()  # 这将隐藏分页按钮

            if choose == 'ALL':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                conditions, dist_stats = statistic_project_status(self.df1, ic_type, factory, glass, flash, year, grade,publisher)

                # 检查DataFrame是否为空
                if dist_stats.empty or len(dist_stats) == 0:
                    messagebox.showinfo("提示", "该筛选条件下项目数量为0")
                    return
                if len(dist_stats.columns) >= 2:
                    project_counts = dist_stats.iloc[:, 1]  # 获取第二列数据
                else:
                    project_counts = dist_stats.iloc[:, 0]  # 如果只有一列，使用第一列  传入的空Series

                if len(project_counts) != 0:
                    plt.close('all')
                    fig = plt.figure(figsize=(6*multiple-0.1, 4.6*multiple-0.05), dpi=100)
                    f_plot = fig.add_subplot(111)  # 划分区域
                    self.canvas = FigureCanvasTkAgg(fig, master=self)
                    self.canvas.get_tk_widget().place(x=0, y=120)  # 放置位置
                    f_plot.clear()  # 刷新
                    ax = sns.barplot(
                        x='Update_Count',
                        y='Project_Count',
                        data=dist_stats,
                        linewidth=1,
                        # color = '#1f77b4'  # 添加颜色参数
                    )
                    # 添加数据标签
                    for p in ax.patches:
                        ax.annotate(
                            f"{int(p.get_height())}",
                            (p.get_x() + p.get_width() / 2., p.get_height()),
                            ha='center', va='center',
                            xytext=(0, 5),
                            textcoords='offset points',
                            fontsize=10,
                            weight='bold'
                        )
                    title = "项目更新次数分布统计\n(" + " | ".join(conditions) + ")" if conditions else "项目更新次数分布统计（ALL）"
                    plt.title(title, fontsize=14, pad=10)
                    plt.xlabel('更新次数', fontsize=12)
                    plt.ylabel('项目数量', fontsize=12)
                    plt.xticks(fontsize=10)
                    plt.yticks(fontsize=10)

                    # 自动调整y轴范围，留出标签空间
                    max_value = max((dist_stats.iloc[:, 1]))
                    plt.yticks(np.arange(0, max_value * 1.15 + 1, step=max(1, int(max_value / 10))), fontsize=10)
                    plt.ylim(0, max_value * 1.15)
                else:
                    messagebox.showinfo("提示", "该筛选条件下项目数量为0")


            elif choose == 'IC型号':
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, ic_counts = statistic_ic_projects(self.df1, factory, glass, flash, year, grade,publisher)
                self.draw_histogram(choose, title_conditions, ic_counts)

            elif choose == '模组厂':
                ic_type = self.ic_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, factory_counts = statistic_module_factory(self.df1, ic_type, glass, flash, year,
                                                                            grade,publisher)
                self.module_data = factory_counts  # 保存模组厂数据
                if int(factory_counts.sum()) > 0:
                    # 计算总页数
                    self.total_pages = (len(factory_counts) + 19) // 20
                    if self.total_pages > 1:
                        self.show_module_page(title_conditions=title_conditions)
                    elif self.total_pages == 1:
                        # self.clear_plot()
                        # 绘制图表
                        fig = plt.figure(figsize=(6*multiple-0.1, 4.6*multiple-0.05), dpi=100)
                        f_plot = fig.add_subplot(111)  # 划分区域
                        self.canvas = FigureCanvasTkAgg(fig, master=self)
                        self.canvas.get_tk_widget().place(x=0, y=120)  # 放置位置
                        f_plot.clear()  # 刷新
                        bars = plt.barh(factory_counts.index, factory_counts.values)

                        max_value = max(factory_counts.values)
                        if max_value > 3:
                            # 添加数据标签
                            for bar in bars:
                                width = bar.get_width()
                                plt.text(width + 0.2,  # x位置：宽度值+0.2的偏移
                                         bar.get_y() + bar.get_height() / 2,  # y位置：居中
                                         f'{int(width)}',
                                         va='center',  # 垂直居中
                                         ha='left',  # 水平左对齐
                                         fontsize=10)
                        else:
                            # 添加数据标签
                            for bar in bars:
                                width = bar.get_width()
                                plt.text(width,  # x位置不偏移
                                         bar.get_y() + bar.get_height() / 2,  # y位置：居中
                                         f'{int(width)}',
                                         va='center',  # 垂直居中
                                         ha='left',  # 水平左对齐
                                         fontsize=10)

                        title = "模组厂项目数量统计"
                        if title_conditions:
                            title += "\n(" + " | ".join(title_conditions) + ")"

                        plt.title(title, fontsize=14, pad=10)
                        plt.ylabel('模组厂', fontsize=12)
                        plt.xlabel('项目数量', fontsize=12)
                        # plt.xticks(rotation=90, fontsize=10)
                        plt.xticks(fontsize=10)
                        plt.yticks(fontsize=10)

                        # 自动调整x轴范围，留出标签空间
                        # max_value = max(factory_counts.values)
                        plt.xticks(np.arange(0, max_value * 1.15 + 1, step=max(1, int(max_value / 10))), fontsize=10)
                        plt.xlim(0, max_value * 1.15)
                elif int(factory_counts.sum()) == 0:
                    messagebox.showinfo("提示", "该筛选条件下项目数量为0")

            elif choose == '玻璃厂':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, glass_counts = statistic_glass_projects(self.df1, ic_type, factory, flash, year,
                                                                          grade,publisher)
                self.draw_histogram(choose, title_conditions, glass_counts)

            elif choose == 'Flash':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, flash_counts = statistic_flash_projects(self.df1, ic_type, factory, glass, year,
                                                                          grade,publisher)
                self.draw_histogram(choose, title_conditions, flash_counts)


            elif choose == '年度':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                grade = self.grade_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, year_counts = statistic_project_by_year(self.df1, ic_type, factory, glass, flash,
                                                                          grade,publisher)
                self.draw_histogram(choose, title_conditions, year_counts)

            elif choose == '等级':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                publisher = self.publisher_dropdown.get()

                title_conditions, grade_counts = statistic_project_by_grade(self.df1, ic_type, factory, glass, flash,
                                                                            year,publisher)
                self.draw_histogram(choose, title_conditions, grade_counts)

            elif choose == '发布人':
                ic_type = self.ic_dropdown.get()
                factory = self.module_dropdown.get()
                glass = self.glass_dropdown.get()
                flash = self.flash_dropdown.get()
                year = self.year_dropdown.get()
                grade = self.grade_dropdown.get()

                title_conditions, publisher_counts = statistic_project_by_Publisher(self.df1, ic_type, factory, glass, flash,
                                                                            year,grade)
                self.draw_histogram(choose, title_conditions, publisher_counts)

            # # 设置网格线
            # f_plot.grid(True, linestyle='--', alpha=0.6)
            # 使直方图填充整个画布
            plt.tight_layout()
            if self.canvas:
                self.canvas.draw()

        except Exception as e:
            messagebox.showerror("错误", f"数据分析时发生错误: {str(e)}")
            self.clear_plot()
            self.total_pages = 1
            self.update_page_buttons()

    def show_module_page(self, title_conditions):
        """
        显示当前页码的模组厂数据
        :return:
        """
        if self.module_data is None:
            return

        start_idx = self.current_page * 20
        end_idx = min((self.current_page + 1) * 20, len(self.module_data))
        current_data = self.module_data[start_idx:end_idx]

        # 创建图形和坐标轴(如果不存在)
        if self.fig is None:
            self.fig = plt.figure(figsize=(6*multiple-0.1, 4.6*multiple-0.4), dpi=100)
            self.ax = self.fig.add_subplot(111)

        # 清除之前的画布
        if self.canvas:
            self.clear_plot()

        self.canvas = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas.get_tk_widget().place(x=0, y=120)
        self.ax.clear()

        fixed_height = 0.6  # 固定高度值
        # 计算间距 - 确保最后一页条目少时也能保持相同高度
        n_bars = len(current_data)
        if n_bars < 20:
            # 调整y轴范围使条形高度保持一致
            self.ax.set_ylim(-0.5, 19.5)  # 保持20个条形的空间

        # 绘制柱状图
        bars = self.ax.barh(
            y=range(n_bars),
            width=current_data.values,
            height=fixed_height)

        # 设置y轴刻度标签为实际模组厂名称
        self.ax.set_yticks(range(n_bars))
        self.ax.set_yticklabels(current_data.index)

        # 添加数据标签
        for bar in bars:
            width = bar.get_width()
            self.ax.text(width + 0.2,
                         bar.get_y() + bar.get_height() / 2,
                         f'{int(width)}',
                         va='center',
                         ha='left',
                         fontsize=10)

        title = f"模组厂项目数量统计"
        self.module_title_conditions = title_conditions
        if self.module_title_conditions:
            title += "\n(" + " | ".join(self.module_title_conditions) + ")"
            # 调整子图位置和标题位置，为多行标题留出空间
            self.fig.subplots_adjust(top=0.86)  # 增加顶部空间
        else:
            self.fig.subplots_adjust(top=0.9)  # 默认顶部空间

        self.ax.set_title(title, fontsize=14, pad=10)
        self.ax.set_ylabel('模组厂', fontsize=12)
        self.ax.set_xlabel('项目数量', fontsize=12)

        # 自动调整x轴范围
        global_max = max(self.module_data.values) * 1.15
        # self.ax.xticks(np.arange(0, global_max * 1.15 + 1, step=max(1, int(global_max / 10))), fontsize=10)
        self.ax.set_xlim(0, global_max)

        # 固定x轴刻度
        max_ticks = 10
        interval = max(1, int(global_max / max_ticks))
        self.ax.set_xticks(range(0, int(global_max) + 1, interval))

        self.ax.tick_params(axis='x', labelsize=10)
        self.ax.tick_params(axis='y', labelsize=10)

        if self.canvas:
            self.canvas.draw()

        # 更新翻页按钮状态
        self.update_page_buttons()

    def update_page_buttons(self):
        """
        更新翻页按钮状态
        :return:
        """
        # 隐藏所有分页控件
        self.page_label.place_forget()
        self.btn_prev.place_forget()
        self.btn_next.place_forget()

        # 只有多于1页且是模组厂分析时才显示分页按钮
        if self.total_pages > 1 and self.analyse_dropdown.get() == '模组厂':
            # 更新文本和状态
            self.page_label.config(text=f"第{self.current_page + 1}页/共{self.total_pages}页")
            self.btn_prev.config(
                command=self.prev_page,
                state=tk.NORMAL if self.current_page > 0 else tk.DISABLED
            )
            self.btn_next.config(
                command=self.next_page,
                state=tk.NORMAL if self.current_page < self.total_pages - 1 else tk.DISABLED
            )

            # 放置控件
            self.page_label.place(x=460 * multiple, y=520 * multiple)
            self.btn_prev.place(x=530 * multiple, y=515 * multiple)
            self.btn_next.place(x=550 * multiple, y=515 * multiple)

    def prev_page(self):
        """
        上一页
        """
        if self.current_page > 0:
            self.current_page -= 1
            self.show_module_page(self.module_title_conditions)

    def next_page(self):
        """
        下一页
        """
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.show_module_page(self.module_title_conditions)

    def clear_plot(self):
        """
        清除直方图
        :return:
        """
        if self.canvas:
            self.canvas.get_tk_widget().destroy()  # 销毁画布
            self.canvas = None


class KSAnalysis(tk.Frame):
    """
    Analyse界面 用于客诉记录分析
    """

    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.create_widgets()
        self.file_path_analyseKS = None
        self.df_KS = None
        self.canvas_KS = None  # 用于保存画布对象

    def create_widgets(self):
        """

        :return:
        """
        # 文件路径显示区域
        self.label_file_pathKS = tk.Label(self, text="文 件 路 径:")
        self.label_file_pathKS.place(x=10, y=10)

        self.text_file_pathKS = tk.Entry(self, state="readonly")
        self.text_file_pathKS.place(x=90, y=10, width=320 * multiple + 20, height=25)

        # 文件选择按钮
        self.btn_open_analyseKS = tk.Button(self, text="...", width=3, bg='Gainsboro', command=self.analyseKS_open_file)
        self.btn_open_analyseKS.place(x=410 * multiple, y=10)

        # 将数据分析及其下拉框框起来
        group_frameKS = tk.LabelFrame(self, bd=2, relief="groove", padx=5, pady=5)
        group_frameKS.place(x=450 * multiple, y=10, width=130* multiple, height=90)
        # 数据分析按键

        self.btn_analyseKS = tk.Button(self, text="客诉统计", command=self.get_xl_to_analyseKS, state=tk.DISABLED)
        self.btn_analyseKS.place(x=465 * multiple, y=15, width=100* multiple)

        self.analyse_valuesKS = ['IC型号', '模组厂', '玻璃厂', '年度', '转接AE与否', '结案与否', '客诉类型', '负责人']
        self.analyse_dropdownKS = ttk.Combobox(self, values=self.analyse_valuesKS, state="readonly")
        self.analyse_dropdownKS.current(0)
        self.analyse_dropdownKS.place(x=465 * multiple, y=60, width=100* multiple)

        self.analyse_dropdownKS.bind("<<ComboboxSelected>>", self.choose_statusKS)

        # 下拉框区域
        self.label_ic_KS = tk.Label(self, text="IC型号:", anchor="e")
        self.label_ic_KS.place(x=10 * multiple, y=50, width=50)
        # self.ic_values_KS = ["ALL", '7272', '7202H', '7202M']
        self.ic_values_KS = ['ALL']
        # self.ic_dropdownKS = ttk.Combobox(self, values=self.ic_values_KS, state="readonly")
        self.ic_dropdownKS = EditableCombobox(self, category='IC', values=self.ic_values_KS)
        # self.ic_dropdownKS.config(state=tk.DISABLED)
        self.ic_dropdownKS.current(0)
        self.ic_dropdownKS.place(x=45 * multiple, y=50, width=90)
        self.ic_dropdownKS.config(state="disabled")

        self.label_module_KS = tk.Label(self, text="模组厂:", anchor="e")
        self.label_module_KS.place(x=120 * multiple, y=50, width=50)
        # self.module_values_KS = ['ALL', '海菲', '信利', '易快来', '联创', '创维', '同兴达', '海盛捷', '壹星', '精卓', '三龙', '合力泰', '华显',
        #                          '维立',
        #                          '亿华', '立德', '晶泰', '德智欣', '中光电', '沛宏', '欣欣光电', '众铭安', '天正达', '瑞恒光电', '清创高', '汉龙时代',
        #                          '晶胜通', '龙煜', '金宏光电', '威达光电', '惠科', '华视', '正金晶光电', '华映', '长信新显', '宏凯',
        #                          '泰启', '百业', '共赢', '德实', '京龙', '如新电子', '皓显', '大通显示', '高展', '康华显通', '煜鑫', '宏利超显', '菲触显视',
        #                          '轩达', '钜沣', '鹰芒技术', '德普特', '元格', '亿普拉斯', '万联', '重联', '日日佳', '中正威', '天山电子', '中正威',
        #                          '天正达', '凯晟', '天山']
        self.module_values_KS = ['ALL']
        # self.module_dropdownKS = ttk.Combobox(self, values=self.module_values_KS, state="readonly")
        self.module_dropdownKS = EditableCombobox(self, category='Module', values=self.module_values_KS)
        self.module_dropdownKS.current(0)
        self.module_dropdownKS.place(x=155 * multiple, y=50, width=90)

        self.label_glass_KS = tk.Label(self, text="玻璃厂:", anchor="e")
        self.label_glass_KS.place(x=230 * multiple, y=50, width=50)
        # self.glass_values_KS = ['ALL', 'CSOT', 'TM', 'TRULY', 'BOE', 'CTO', 'PANDA', 'SHARP', 'MDT', 'HKC', 'HSD',
        #                         'INX',
        #                         'IVO', 'CTC']
        self.glass_values_KS = ['ALL']
        # self.glass_dropdownKS = ttk.Combobox(self, values=self.glass_values_KS, state="readonly")
        self.glass_dropdownKS = EditableCombobox(self, category='Glass', values=self.glass_values_KS)
        self.glass_dropdownKS.current(0)
        self.glass_dropdownKS.place(x=265 * multiple, y=50, width=90)

        self.label_year_KS = tk.Label(self, text="年度:", anchor="e")
        self.label_year_KS.place(x=335 * multiple, y=50, width=50)
        # self.year_values_KS = ['ALL', '2022', '2023', '2024', '2025']
        self.year_values_KS = ['ALL']
        # self.year_dropdownKS = ttk.Combobox(self, values=self.year_values_KS, state="readonly")
        self.year_dropdownKS = EditableCombobox(self, category='Year', values=self.year_values_KS)
        self.year_dropdownKS.current(0)
        self.year_dropdownKS.place(x=370 * multiple, y=50, width=90)

        self.label_AE_KS = tk.Label(self, text="转接AE与否:", anchor="e")
        self.label_AE_KS.place(x=1, y=80, width=70)
        self.AE_values_KS = ['ALL', '是', '否']
        self.AE_dropdownKS = ttk.Combobox(self, values=self.AE_values_KS, state="readonly")
        self.AE_dropdownKS.current(0)
        self.AE_dropdownKS.place(x=45 * multiple, y=80, width=90)

        self.label_over_KS = tk.Label(self, text="结案与否:", anchor="e")
        self.label_over_KS.place(x=115 * multiple, y=80, width=60)
        self.over_values_KS = ['ALL', '是', '否']
        self.over_dropdownKS = ttk.Combobox(self, values=self.over_values_KS, state="readonly")
        self.over_dropdownKS.current(0)
        self.over_dropdownKS.place(x=155 * multiple, y=80, width=90)

        self.label_type_KS = tk.Label(self, text="客诉类型:", anchor="e")
        self.label_type_KS.place(x=225 * multiple, y=80, width=60)
        # self.type_values_KS = ['ALL', '误判', 'DPcode', 'FW效果', 'IC来料', '环境干扰', '模组工艺', '人为因素', '玻璃']
        self.type_values_KS = ['ALL']
        # self.type_dropdownKS = ttk.Combobox(self, values=self.type_values_KS, state="readonly")
        self.type_dropdownKS = EditableCombobox(self, category='Type', values=self.type_values_KS)
        self.type_dropdownKS.current(0)
        self.type_dropdownKS.place(x=265 * multiple, y=80, width=90)

        self.label_principal_KS = tk.Label(self, text="负责人:", anchor="e")
        self.label_principal_KS.place(x=335 * multiple, y=80, width=50)
        # self.principal_values_KS = ['ALL', '黄耀', '吴宏博', '丁淑芳', '成鹏', '李尚聪', '卓秀慧']
        self.principal_values_KS = ['ALL']
        # self.principal_dropdownKS = ttk.Combobox(self, values=self.principal_values_KS, state="readonly")
        self.principal_dropdownKS = EditableCombobox(self, category='Principal', values=self.principal_values_KS)
        self.principal_dropdownKS.current(0)
        self.principal_dropdownKS.place(x=370 * multiple, y=80, width=90)

    def analyseKS_open_file(self):
        """

        :return:
        """
        # 打开文件选择对话框
        self.file_path_analyseKS = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])

        if self.file_path_analyseKS:
            # 更新文件路径
            self.text_file_pathKS.config(state="normal")
            self.text_file_pathKS.delete(0, tk.END)
            self.text_file_pathKS.insert(0, self.file_path_analyseKS)
            self.text_file_pathKS.config(state="readonly")

            # 启用数据分析键
            self.btn_analyseKS.config(state=tk.NORMAL)

            # 重新选择文件则清除画布
            # self.clear_plot()

    def analyseKS_reload_excel(self):
        """
        加载文件
        :return:
        """
        try:
            self.df_KS = read_excel_for_analyseKS(self.file_path_analyseKS, sheet_name='客诉记录')
        except Exception as e:
            self.df_KS = None
            messagebox.showerror("错误", f"读取Excel文件失败: {e}")

    def choose_statusKS(self, event=None):
        choose = self.analyse_dropdownKS.get()
        if choose == 'IC型号':
            self.ic_dropdownKS.config(state=tk.DISABLED)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '模组厂':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.DISABLED)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            # self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '玻璃厂':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.DISABLED)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            # self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '年度':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.DISABLED)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            # self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '转接AE与否':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.DISABLED)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            # self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '结案与否':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.DISABLED)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            # self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '客诉类型':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.DISABLED)
            self.principal_dropdownKS.config(state=tk.NORMAL)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            # self.type_dropdownKS.config(state="readonly")
            self.principal_dropdownKS.config(state="readonly")

        elif choose == '负责人':
            self.ic_dropdownKS.config(state=tk.NORMAL)
            self.module_dropdownKS.config(state=tk.NORMAL)
            self.glass_dropdownKS.config(state=tk.NORMAL)
            self.year_dropdownKS.config(state=tk.NORMAL)
            self.AE_dropdownKS.config(state=tk.NORMAL)
            self.over_dropdownKS.config(state=tk.NORMAL)
            self.type_dropdownKS.config(state=tk.NORMAL)
            self.principal_dropdownKS.config(state=tk.DISABLED)
            self.ic_dropdownKS.config(state="readonly")
            self.module_dropdownKS.config(state="readonly")
            self.glass_dropdownKS.config(state="readonly")
            self.year_dropdownKS.config(state="readonly")
            self.AE_dropdownKS.config(state="readonly")
            self.over_dropdownKS.config(state="readonly")
            self.type_dropdownKS.config(state="readonly")
            # self.principal_dropdownKS.config(state="readonly")

    def draw_histogram_KS(self, choose, title, counts):
        """
        绘制直方图
        :param choose: X粥
        :param title: 筛选条件
        :param counts: 直方图数据
        :return: 直方图
        """
        # 绘制图表
        if int(counts.sum()) != 0:
            fig = plt.figure(figsize=(6*multiple-0.1, 4.6*multiple-0.05), dpi=100)
            f_plot = fig.add_subplot(111)  # 划分区域
            self.canvas = FigureCanvasTkAgg(fig, master=self)
            self.canvas.get_tk_widget().place(x=0, y=120)  # 放置位置
            f_plot.clear()  # 刷新

            bars = plt.bar(counts.index, counts.values, width=0.6)
            # 添加数据标签
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{int(height)}',
                         ha='center', va='bottom', fontsize=11, weight='bold')

            a_title = choose + "客诉数量统计"
            if title:
                a_title += "\n(" + " | ".join(title) + ")"

            plt.title(a_title, fontsize=14, pad=10)
            plt.xlabel(choose, fontsize=12)
            plt.ylabel('客诉记录数量', fontsize=12)
            if choose == '模组厂':
                plt.xticks(rotation=90, fontsize=10)
            else:
                plt.xticks(fontsize=10)
            # plt.yticks(fontsize=10)

            # 自动调整y轴范围，留出标签空间
            max_value = max(counts.values)
            plt.yticks(np.arange(0, max_value * 1.15 + 1, step=max(1, int(max_value / 10))), fontsize=10)
            plt.ylim(0, max_value * 1.15)
        else:
            messagebox.showinfo("提示", "该筛选条件下项目数量为0")

    def get_xl_to_analyseKS(self):
        """
        根据不同选择情况绘制直方图
        :return:
        """
        try:
            self.analyseKS_reload_excel()  # 每次点击数据分析可重新加载文件
            choose = self.analyse_dropdownKS.get()
            self.clear_plotKS()  # 每次选择先清除上一次直方图

            if choose == 'IC型号':
                factory = self.module_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                year = self.year_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()
                title_conditions, ic_counts = statistic_ic_projects_KS(self.df_KS, factory, glass, year, transfer_AE,
                                                                       over, type, principal)
                self.draw_histogram_KS(choose, title_conditions, ic_counts)


            elif choose == '模组厂':
                ic_type = self.ic_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                year = self.year_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, factory_counts = statistic_module_factory_KS(self.df_KS, ic_type, glass, year,
                                                                               transfer_AE, over, type, principal)
                self.draw_histogram_KS(choose, title_conditions, factory_counts)

            elif choose == '玻璃厂':
                factory = self.module_dropdownKS.get()
                ic_type = self.ic_dropdownKS.get()
                year = self.year_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, glass_counts = statistic_glass_projects_KS(self.df_KS, factory, ic_type, year,
                                                                             transfer_AE, over, type, principal)
                self.draw_histogram_KS(choose, title_conditions, glass_counts)

            elif choose == '年度':
                ic_type = self.ic_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                factory = self.module_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, year_counts = statistic_project_by_year_KS(self.df_KS, factory, ic_type, glass,
                                                                             transfer_AE, over, type, principal)
                self.draw_histogram_KS(choose, title_conditions, year_counts)

            elif choose == '转接AE与否':
                ic_type = self.ic_dropdownKS.get()
                factory = self.module_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                year = self.year_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, transferAE_counts = statistic_project_transferAE_KS(self.df_KS, factory, ic_type,
                                                                                      glass, year, over, type,
                                                                                      principal)
                self.draw_histogram_KS(choose, title_conditions, transferAE_counts)

            elif choose == '结案与否':
                factory = self.module_dropdownKS.get()
                ic_type = self.ic_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                year = self.year_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                type = self.type_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, over_counts = statistic_project_over_KS(self.df_KS, factory, ic_type, glass, year,
                                                                          transfer_AE, type, principal)
                self.draw_histogram_KS(choose, title_conditions, over_counts)

            elif choose == '客诉类型':
                factory = self.module_dropdownKS.get()
                ic_type = self.ic_dropdownKS.get()
                year = self.year_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                principal = self.principal_dropdownKS.get()

                title_conditions, type_counts = statistic_project_type_KS(self.df_KS, factory, ic_type, glass, year,
                                                                          transfer_AE, over, principal)
                self.draw_histogram_KS(choose, title_conditions, type_counts)

            elif choose == '负责人':
                ic_type = self.ic_dropdownKS.get()
                glass = self.glass_dropdownKS.get()
                factory = self.module_dropdownKS.get()
                year = self.year_dropdownKS.get()
                transfer_AE = self.AE_dropdownKS.get()
                over = self.over_dropdownKS.get()
                type = self.type_dropdownKS.get()

                title_conditions, principal_counts = statistic_project_principal_KS(self.df_KS, factory, ic_type, glass,
                                                                                    year, transfer_AE, over, type)
                self.draw_histogram_KS(choose, title_conditions, principal_counts)
            # # 设置网格线
            # f_plot.grid(True, linestyle='--', alpha=0.6)
            # 使直方图填充整个画布
            plt.tight_layout()
            if self.canvas_KS:
                self.canvas_KS.draw()

        except Exception as e:
            messagebox.showerror("错误", f"数据分析时发生错误: {str(e)}")
            self.clear_plotKS()

    def clear_plotKS(self):
        """
        清除直方图
        :return:
        """
        if self.canvas_KS:
            self.canvas_KS.get_tk_widget().destroy()  # 销毁画布
            self.canvas_KS = None



class HNAnalysis(tk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.create_widgets()
        self.file_path_analyse_HN = None
        self.df_HN = None
        self.canvas_HN = None  # 用于保存画布对象
        self.current_page_HN = 0
        self.total_pages_HN = 1
        self.module_data_HN = None
        self.module_title_conditions_HN = None
        self.counts_sum = None
        self.animation_HN = None  # 用于存储动画对象
        self.fig_HN = None  # 用于存储图形对象
        self.ax_HN = None  # 用于存储坐标轴对象

    def create_widgets(self):
        """
        创建界面组件
        """
        # 文件路径显示区域
        self.label1_file_path_HN = tk.Label(self, text="文 件 路 径:")
        self.label1_file_path_HN.place(x=10, y=10)

        self.text1_file_path_HN = tk.Entry(self, state="readonly")
        self.text1_file_path_HN.place(x=90, y=10, width=320 * multiple + 30, height=25)

        # 文件选择按钮
        self.btn_open_analyse_HN = tk.Button(self, text="...", width=3, bg='Gainsboro',
                                             command=self.analyse_open_file_HN)
        self.btn_open_analyse_HN.place(x=410* multiple, y=10)

        # 将数据分析及其下拉框框起来
        group_frame_HN = tk.LabelFrame(self, bd=2, relief="groove", padx=5, pady=5)
        group_frame_HN.place(x=450* multiple, y=10, width=130* multiple, height=90)

        # 数据分析按键
        self.btn_analyse_HN = tk.Button(self, text="项目统计", command=self.get_xl_to_analyse_HN,
                                        state=tk.DISABLED)
        self.btn_analyse_HN.place(x=465* multiple, y=15 ,width=100* multiple)

        self.analyse_values_HN = ['IC型号', '模组厂', '玻璃型号', '年度','立项时间', 'TP Mode', '贴脸息屏', '项目状态', '类别', '销售', 'PM']
        self.analyse_dropdown_HN = ttk.Combobox(self, values=self.analyse_values_HN, state="readonly")
        self.analyse_dropdown_HN.current(0)
        self.analyse_dropdown_HN.place(x=465* multiple, y=60, width=100* multiple)

        self.analyse_dropdown_HN.bind("<<ComboboxSelected>>", self.choose_status_HN)

        # 下拉框区域
        self.label_ic_HN = tk.Label(self, text="IC型号:", anchor="e")
        self.label_ic_HN.place(x=10* multiple, y=50, width=50)
        self.ic_values_HN = ['ALL','Unknown']
        self.ic_dropdown_HN = EditableCombobox(self, category='IC_Type', sheet_name='Sheet3', values=self.ic_values_HN)
        self.ic_dropdown_HN.current(0)
        self.ic_dropdown_HN.place(x=45* multiple, y=50, width=90)
        self.ic_dropdown_HN.config(state="disabled")

        self.label_module_HN = tk.Label(self, text="模组厂:", anchor="e")
        self.label_module_HN.place(x=120* multiple, y=50, width=50)
        self.module_values_HN = ['ALL','Unknown']
        self.module_dropdown_HN = EditableCombobox(self, category='Factory', sheet_name='Sheet3',
                                                   values=self.module_values_HN)
        self.module_dropdown_HN.current(0)
        self.module_dropdown_HN.place(x=155* multiple, y=50, width=90)

        self.label_glass_HN = tk.Label(self, text="玻璃型号:", anchor="e")
        self.label_glass_HN.place(x=225* multiple, y=50, width=60)
        self.glass_values_HN = ['ALL','Unknown']
        self.glass_dropdown_HN = EditableCombobox(self, category='Glass', sheet_name='Sheet3',
                                                  values=self.glass_values_HN)
        self.glass_dropdown_HN.current(0)
        self.glass_dropdown_HN.place(x=265* multiple, y=50, width=90)

        self.label_year_HN = tk.Label(self, text="年度:", anchor="e")
        self.label_year_HN.place(x=335* multiple, y=50, width=50)
        self.year_values_HN = ['ALL']
        self.year_dropdown_HN = EditableCombobox(self, category='Year', sheet_name='Sheet3', values=self.year_values_HN)
        self.year_dropdown_HN.current(0)
        self.year_dropdown_HN.place(x=370* multiple, y=50, width=90)

        self.label_month_HN = tk.Label(self, text="立项时间:", anchor="e")
        self.label_month_HN.place(x=5* multiple, y=80, width=60)
        self.month_values_HN = ['ALL']
        self.month_dropdown_HN = EditableCombobox(self, category='Month', sheet_name='Sheet3',
                                                  values=self.month_values_HN)
        self.month_dropdown_HN.current(0)
        self.month_dropdown_HN.place(x=45* multiple, y=80, width=90)

        self.label_tpmode_HN = tk.Label(self, text="TP Mode:", anchor="e")
        self.label_tpmode_HN.place(x=114* multiple, y=80, width=60)
        self.tpmode_values_HN = ['ALL', 'NULL','Unknown']
        self.tpmode_dropdown_HN = EditableCombobox(self, category='TP_Mode', sheet_name='Sheet3',
                                                   values=self.tpmode_values_HN)
        self.tpmode_dropdown_HN.current(0)
        self.tpmode_dropdown_HN.place(x=155* multiple, y=80, width=90)

        self.label_palm_HN = tk.Label(self, text="贴脸息屏:", anchor="e")
        self.label_palm_HN.place(x=225* multiple, y=80, width=60)
        self.palm_values_HN = ['ALL', 'NULL','Unknown']
        self.palm_dropdown_HN = EditableCombobox(self, category='Palm', sheet_name='Sheet3', values=self.palm_values_HN)
        self.palm_dropdown_HN.current(0)
        self.palm_dropdown_HN.place(x=265* multiple, y=80, width=90)

        self.label_status_HN = tk.Label(self, text="项目状态:", anchor="e")
        self.label_status_HN.place(x=330* multiple, y=80, width=60)
        self.status_values_HN = ['ALL','Unknown']
        self.status_dropdown_HN = EditableCombobox(self, category='Status', sheet_name='Sheet3',
                                                   values=self.status_values_HN)
        self.status_dropdown_HN.current(0)
        self.status_dropdown_HN.place(x=370* multiple, y=80, width=90)


        self.label_category_HN = tk.Label(self, text="类别:", anchor="e")
        self.label_category_HN.place(x=12* multiple, y=110, width=50)
        self.category_values_HN = ['ALL', 'NULL','Unknown']
        self.category_dropdown_HN = EditableCombobox(self, category='Category', sheet_name='Sheet3',
                                                     values=self.category_values_HN)
        self.category_dropdown_HN.current(0)
        self.category_dropdown_HN.place(x=45* multiple, y=110, width=90)

        self.label_sales_HN = tk.Label(self, text="销售:", anchor="e")
        self.label_sales_HN.place(x=120* multiple, y=110, width=50)
        self.sales_values_HN = ['ALL']
        self.sales_dropdown_HN = EditableCombobox(self, category='Sales', sheet_name='Sheet3',
                                                  values=self.sales_values_HN)
        self.sales_dropdown_HN.current(0)
        self.sales_dropdown_HN.place(x=155* multiple, y=110, width=90)

        self.label_pm_HN = tk.Label(self, text="PM:", anchor="e")
        self.label_pm_HN.place(x=232* multiple, y=110, width=50)
        self.pm_values_HN = ['ALL']
        self.pm_dropdown_HN = EditableCombobox(self, category='PM', sheet_name='Sheet3', values=self.pm_values_HN)
        self.pm_dropdown_HN.current(0)
        self.pm_dropdown_HN.place(x=265* multiple, y=110, width=90)

        # 初始化分页控件
        self.page_label_HN = tk.Label(self, text="", font=('Arial', 10))
        self.btn_prev_HN = tk.Button(self, text="◀", width=2, command=self.prev_page_HN)
        self.btn_next_HN = tk.Button(self, text="▶", width=2, command=self.next_page_HN)

        # 初始隐藏分页控件
        self.page_label_HN.place_forget()
        self.btn_prev_HN.place_forget()
        self.btn_next_HN.place_forget()

    def analyse_open_file_HN(self):
        """
        打开文件选择对话框
        """
        self.file_path_analyse_HN = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx")])

        if self.file_path_analyse_HN:
            # 更新文件路径
            self.text1_file_path_HN.config(state="normal")
            self.text1_file_path_HN.delete(0, tk.END)
            self.text1_file_path_HN.insert(0, self.file_path_analyse_HN)
            self.text1_file_path_HN.config(state="readonly")

            # 启用数据分析键
            self.btn_analyse_HN.config(state=tk.NORMAL)

            # 重新选择文件则清除画布
            self.clear_plot_HN()

    def analyse_reload_excel_HN(self):
        """
        加载文件
        """
        try:
            file_name = self.file_path_analyse_HN
            self.df_HN = read_excel_for_analyseHN(file_name)
            print(f"成功加载文件，数据形状: {self.df_HN.shape}")
        except Exception as e:
            self.df_HN = None
            messagebox.showerror("错误", f"读取Excel文件失败: {e}")

    def choose_status_HN(self, event=None):
        """
        根据选择的分析类型启用/禁用对应的筛选条件
        """
        choose = self.analyse_dropdown_HN.get()

        # 重置所有下拉框为可编辑状态
        self.ic_dropdown_HN.config(state="readonly")
        self.module_dropdown_HN.config(state="readonly")
        self.glass_dropdown_HN.config(state="readonly")
        self.year_dropdown_HN.config(state="readonly")
        self.month_dropdown_HN.config(state="readonly")
        self.tpmode_dropdown_HN.config(state="readonly")
        self.palm_dropdown_HN.config(state="readonly")
        self.status_dropdown_HN.config(state="readonly")
        self.category_dropdown_HN.config(state="readonly")
        self.sales_dropdown_HN.config(state="readonly")
        self.pm_dropdown_HN.config(state="readonly")

        # 根据选择禁用对应的下拉框
        if choose == 'IC型号':
            self.ic_dropdown_HN.config(state="disabled")
        elif choose == '模组厂':
            self.module_dropdown_HN.config(state="disabled")
        elif choose == '立项时间':
            self.month_dropdown_HN.config(state="disabled")
        elif choose == '年度':
            self.year_dropdown_HN.config(state="disabled")
        elif choose == '玻璃型号':
            self.glass_dropdown_HN.config(state="disabled")
        elif choose == 'TP Mode':
            self.tpmode_dropdown_HN.config(state="disabled")
        elif choose == '贴脸息屏':
            self.palm_dropdown_HN.config(state="disabled")
        elif choose == '项目状态':
            self.status_dropdown_HN.config(state="disabled")
        elif choose == '类别':
            self.category_dropdown_HN.config(state="disabled")
        elif choose == '销售':
            self.sales_dropdown_HN.config(state="disabled")
        elif choose == 'PM':
            self.pm_dropdown_HN.config(state="disabled")

    def draw_histogram_HN(self, choose, title, counts, counts_sum):
        """
        绘制直方图
        :param choose: X轴
        :param title: 筛选条件
        :param counts: 直方图数据
        """
        # 绘制图表
        if int(counts.sum()) != 0:
            fig_HN = plt.figure(figsize=(6*multiple-0.1, 4.4*multiple-0.05), dpi=100)
            f_plot_HN = fig_HN.add_subplot(111)  # 划分区域
            self.canvas_HN = FigureCanvasTkAgg(fig_HN, master=self)
            self.canvas_HN.get_tk_widget().place(x=0, y=150)  # 放置位置
            f_plot_HN.clear()  # 刷新

            bars = plt.bar(counts.index, counts.values, width=0.6)
            # 添加数据标签
            for bar in bars:
                height = bar.get_height()
                plt.text(bar.get_x() + bar.get_width() / 2., height,
                         f'{int(height)}',
                         ha='center', va='bottom', fontsize=11, weight='bold')

            a_title = f"{choose}华南项目统计 (项目总数:{int(counts_sum)})"
            if title:
                a_title += "\n(" + " | ".join(title) + ")"

            plt.title(a_title, fontsize=14, pad=10)
            plt.xlabel(choose, fontsize=12)
            plt.ylabel('项目数量', fontsize=12)
            if choose == '模组厂':
                plt.xticks(rotation=90, fontsize=10)
            else:
                plt.xticks(fontsize=10)

            # 自动调整y轴范围，留出标签空间
            max_value = max(counts.values)
            plt.yticks(np.arange(0, max_value * 1.15 + 1, step=max(1, int(max_value / 10))), fontsize=10)
            plt.ylim(0, max_value * 1.15)

            # 使直方图填充整个画布
            plt.tight_layout()
            if self.canvas_HN:
                self.canvas_HN.draw()
            # 隐藏分页按钮，因为非分页类型不需要
            self.page_label_HN.place_forget()
            self.btn_prev_HN.place_forget()
            self.btn_next_HN.place_forget()

        else:
            messagebox.showinfo("提示", "该筛选条件下项目数量为0")

    def get_xl_to_analyse_HN(self):
        """
        根据不同选择情况绘制直方图
        """
        try:
            self.analyse_reload_excel_HN()  # 每次点击数据分析可重新加载文件
            choose = self.analyse_dropdown_HN.get()
            self.clear_plot_HN()  # 每次选择先清除上一次直方图

            # 重置分页状态
            self.current_page_HN = 0
            self.total_pages_HN = 1

            # 获取筛选条件
            ic_type = self.ic_dropdown_HN.get()
            factory = self.module_dropdown_HN.get()
            glass = self.glass_dropdown_HN.get()
            year = self.year_dropdown_HN.get()
            month = self.month_dropdown_HN.get()
            tpmode = self.tpmode_dropdown_HN.get()
            palm = self.palm_dropdown_HN.get()
            status = self.status_dropdown_HN.get()
            category = self.category_dropdown_HN.get()
            sales = self.sales_dropdown_HN.get()
            pm = self.pm_dropdown_HN.get()

            if choose == 'IC型号':
                title_conditions, ic_counts,ic_counts_sum = statistic_ic_projects_HN(
                    self.df_HN, factory, glass, year, month, tpmode, palm,
                    status, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, ic_counts,ic_counts_sum)

            elif choose == '模组厂':
                title_conditions, factory_counts,factory_counts_sum = statistic_factory_projects_HN(
                    self.df_HN, ic_type, glass, year, month, tpmode, palm,
                    status, category, sales, pm
                )
                self.module_data_HN = factory_counts  # 保存模组厂数据
                if int(factory_counts.sum()) > 0:
                    # 计算总页数
                    self.total_pages_HN = (len(factory_counts) + 19) // 20
                    self.counts_sum = factory_counts_sum
                    if self.total_pages_HN > 1:
                        self.show_module_page_HN(title_conditions=title_conditions, choose=choose,counts_sum=factory_counts_sum)
                    elif self.total_pages_HN == 1:
                        # 使用分页显示函数显示单页数据
                        self.show_module_page_HN(title_conditions=title_conditions, choose=choose,counts_sum=factory_counts_sum)
                elif int(factory_counts.sum()) == 0:
                    messagebox.showinfo("提示", "该筛选条件下项目数量为0")

            elif choose == '立项时间':
                title_conditions, month_counts, month_counts_sum = statistic_month_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, tpmode, palm,
                    status, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, month_counts,month_counts_sum)

            elif choose == '年度':
                title_conditions, year_counts,year_counts_sum = statistic_year_projects_HN(
                    self.df_HN, ic_type, factory, glass, month, tpmode, palm,
                    status, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, year_counts, year_counts_sum)

            elif choose == '玻璃型号':
                title_conditions, glass_counts, glass_counts_sum= statistic_glass_projects_HN(
                    self.df_HN, ic_type, factory, year, month, tpmode, palm,
                    status, category, sales, pm
                )
                self.module_data_HN = glass_counts  # 保存玻璃型号数据
                if int(glass_counts.sum()) > 0:
                    # 计算总页数
                    self.total_pages_HN = (len(glass_counts) + 19) // 20
                    self.counts_sum = glass_counts_sum
                    if self.total_pages_HN > 1:
                        self.show_module_page_HN(title_conditions=title_conditions, choose=choose,counts_sum=glass_counts_sum)
                    elif self.total_pages_HN == 1:
                        # 使用分页显示函数显示单页数据
                        self.show_module_page_HN(title_conditions=title_conditions, choose=choose,counts_sum=glass_counts_sum)
                elif int(glass_counts.sum()) == 0:
                    messagebox.showinfo("提示", "该筛选条件下项目数量为0")

            elif choose == 'TP Mode':
                title_conditions, tpmode_counts,tpmode_counts_sum = statistic_tpmode_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, palm,
                    status, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, tpmode_counts, tpmode_counts_sum)

            elif choose == '贴脸息屏':
                title_conditions, palm_counts,  palm_counts_sum= statistic_palm_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, tpmode,
                    status, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, palm_counts,palm_counts_sum)

            elif choose == '项目状态':
                title_conditions, status_counts,status_counts_sum = statistic_status_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, tpmode,
                    palm, category, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, status_counts,status_counts_sum)

            elif choose == '类别':
                title_conditions, category_counts,category_counts_sum  = statistic_category_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, tpmode,
                    palm, status, sales, pm
                )
                self.draw_histogram_HN(choose, title_conditions, category_counts,category_counts_sum)

            elif choose == '销售':
                title_conditions, sales_counts,sales_counts_sum = statistic_sales_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, tpmode,
                    palm, status, category, pm
                )
                self.draw_histogram_HN(choose, title_conditions, sales_counts, sales_counts_sum)

            elif choose == 'PM':
                title_conditions, pm_counts, pm_counts_sum= statistic_pm_projects_HN(
                    self.df_HN, ic_type, factory, glass, year, month, tpmode,
                    palm, status, category, sales
                )
                self.draw_histogram_HN(choose, title_conditions, pm_counts, pm_counts_sum)

            # 使直方图填充整个画布
            plt.tight_layout()
            if self.canvas_HN:
                self.canvas_HN.draw()

            # 更新翻页按钮状态
            self.update_page_buttons_HN(choose)

        except Exception as e:
            messagebox.showerror("错误", f"数据分析时发生错误: {str(e)}")
            self.clear_plot_HN()
            self.total_pages_HN = 1
            self.update_page_buttons_HN(choose)

    def show_module_page_HN(self, title_conditions, choose, counts_sum):
        """
        显示当前页码的模组厂或玻璃型号数据
        :param title_conditions: 筛选条件
        :param choose: 当前选择的统计类型
        """
        if self.module_data_HN is None:
            return

        start_idx = self.current_page_HN * 20
        end_idx = min((self.current_page_HN + 1) * 20, len(self.module_data_HN))
        current_data = self.module_data_HN[start_idx:end_idx]

        # 确保先清除之前的画布
        self.clear_plot_HN()

        # 创建图形和坐标轴(如果不存在)
        if self.fig_HN is None:
            self.fig_HN = plt.figure(figsize=(6*multiple-0.1, 4.4*multiple-0.4), dpi=100)
            self.ax_HN = self.fig_HN.add_subplot(111)

        # 清除之前的画布
        # if self.canvas_HN:
        #     self.clear_plot_HN()

        self.canvas_HN = FigureCanvasTkAgg(self.fig_HN, master=self)
        self.canvas_HN.get_tk_widget().place(x=0, y=150)
        self.ax_HN.clear()

        fixed_height = 0.6  # 固定高度值
        # 计算间距 - 确保最后一页条目少时也能保持相同高度
        n_bars = len(current_data)
        if n_bars < 20:
            # 调整y轴范围使条形高度保持一致
            self.ax_HN.set_ylim(-0.5, 19.5)  # 保持20个条形的空间

        # 绘制柱状图
        bars = self.ax_HN.barh(
            y=range(n_bars),
            width=current_data.values,
            height=fixed_height)

        # 设置y轴刻度标签为实际名称
        self.ax_HN.set_yticks(range(n_bars))
        self.ax_HN.set_yticklabels(current_data.index)

        # 添加数据标签
        for bar in bars:
            width = bar.get_width()
            self.ax_HN.text(width + 0.2,
                            bar.get_y() + bar.get_height() / 2,
                            f'{int(width)}',
                            va='center',
                            ha='left',
                            fontsize=10)

        title = f"{choose}华南项目统计 (项目总数:{int(counts_sum)})"
        self.module_title_conditions_HN = title_conditions
        if self.module_title_conditions_HN:
            title += "\n(" + " | ".join(self.module_title_conditions_HN) + ")"
            # 调整子图位置和标题位置，为多行标题留出空间
            self.fig_HN.subplots_adjust(top=0.86)  # 增加顶部空间
        else:
            self.fig_HN.subplots_adjust(top=0.9)  # 默认顶部空间

        self.ax_HN.set_title(title, fontsize=14, pad=10)
        self.ax_HN.set_ylabel(choose, fontsize=12)
        self.ax_HN.set_xlabel('项目数量', fontsize=12)

        # 自动调整x轴范围
        global_max = max(self.module_data_HN.values) * 1.15
        self.ax_HN.set_xlim(0, global_max)

        # 固定x轴刻度
        max_ticks = 10
        interval = max(1, int(global_max / max_ticks))
        self.ax_HN.set_xticks(range(0, int(global_max) + 1, interval))

        self.ax_HN.tick_params(axis='x', labelsize=10)
        self.ax_HN.tick_params(axis='y', labelsize=10)

        # if self.canvas_HN:
        #     self.canvas_HN.draw()
        plt.tight_layout()
        self.canvas_HN.draw()

        #更新翻页按键状态
        self.update_page_buttons_HN(choose)

    def update_page_buttons_HN(self, choose):
        """
        更新翻页按钮状态
        :param choose: 当前选择的统计类型
        """
        # 隐藏所有分页控件
        self.page_label_HN.place_forget()
        self.btn_prev_HN.place_forget()
        self.btn_next_HN.place_forget()

        # 只有多于1页且是模组厂或玻璃型号分析时才显示分页按钮
        if self.total_pages_HN > 0 and choose in ['模组厂', '玻璃型号']:
            # 更新文本和状态
            self.page_label_HN.config(text=f"第{self.current_page_HN + 1}页/共{self.total_pages_HN}页")
            self.btn_prev_HN.config(
                state=tk.NORMAL if self.current_page_HN > 0 else tk.DISABLED
            )
            self.btn_next_HN.config(
                state=tk.NORMAL if self.current_page_HN < self.total_pages_HN - 1 else tk.DISABLED
            )

            # 放置控件
            self.page_label_HN.place(x=460* multiple, y=520* multiple)
            self.btn_prev_HN.place(x=530* multiple, y=515* multiple)
            self.btn_next_HN.place(x=550* multiple, y=515* multiple)

    def prev_page_HN(self):
        """
        上一页
        """
        if self.current_page_HN > 0:
            self.current_page_HN -= 1
            choose = self.analyse_dropdown_HN.get()
            self.show_module_page_HN(self.module_title_conditions_HN, choose,self.counts_sum)

    def next_page_HN(self):
        """
        下一页
        """
        if self.current_page_HN < self.total_pages_HN - 1:
            self.current_page_HN += 1
            choose = self.analyse_dropdown_HN.get()
            self.show_module_page_HN(self.module_title_conditions_HN, choose,self.counts_sum)

    def clear_plot_HN(self):
        """
        清除直方图
        """
        if self.canvas_HN:
            self.canvas_HN.get_tk_widget().destroy()  # 销毁画布
            self.canvas_HN = None
            # 清除图形对象
        if self.fig_HN:
            plt.close(self.fig_HN)
            self.fig_HN = None
            self.ax_HN = None

        # 隐藏分页控件
        self.page_label_HN.place_forget()
        self.btn_prev_HN.place_forget()
        self.btn_next_HN.place_forget()

class MainApp:
    """
    主程序界面，用于切换check和 Analyse界面
    """

    def __init__(self, root):
        self.root = root
        self.root.title("MP List Check Tool_v5.4_20260130")
        # self.root.geometry("600x580")  # 设置窗口大小
        self.root.resizable(0, 0)

        # 创建标签页容器
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True)

        # 创建Check标签页
        self.check_frame = CheckFrame(self.notebook)
        self.notebook.add(self.check_frame, text='   Check   ')

        # 创建Analyse标签页
        self.analyse_frame = AnalyseFrame(self.notebook)
        self.notebook.add(self.analyse_frame, text='  Analyse  ')

        # 绑定标签页切换事件
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

        # 配置标签样式
        self.style = ttk.Style()
        self.style.configure('TNotebook', background='#f0f0f0')
        self.style.configure('TNotebook.Tab',
                             font=('Arial', 10, 'bold'),
                             padding=(20, 5),
                             background='#e0e0e0',
                             foreground='#444')
        self.style.map('TNotebook.Tab',
                       background=[('selected', '#ffffff')],
                       expand=[('selected', [1, 1, 1, 0])])

    def on_tab_changed(self, event):
        """
        标签页切换事件处理
        :return:
        """
        selected_tab = self.notebook.select()
        tab_text = self.notebook.tab(selected_tab, "text")

        if tab_text.strip() == "Analyse":
            # 切换到Analyse界面时，确保直方图显示
            current_tab = self.analyse_frame.notebook.select()
            current_tab_text = self.analyse_frame.notebook.tab(current_tab, "text")

            if current_tab_text.strip() == "MP分析":
                if hasattr(self.analyse_frame.mp_analysis, 'canvas') and self.analyse_frame.mp_analysis.canvas:
                    self.analyse_frame.mp_analysis.canvas.get_tk_widget().place(x=0, y=120)
            elif current_tab_text.strip() == "客诉分析":
                if hasattr(self.analyse_frame.ks_analysis, 'canvas_KS') and self.analyse_frame.ks_analysis.canvas_KS:
                    self.analyse_frame.ks_analysis.canvas_KS.get_tk_widget().place(x=0, y=120)
            elif current_tab_text.strip() == "华南项目":
                if hasattr(self.analyse_frame.HN_analysis, 'canvas_HN') and self.analyse_frame.HN_analysis.canvas_HN:
                    self.analyse_frame.HN_analysis.canvas_HN.get_tk_widget().place(x=0, y=150)

        elif tab_text.strip() == "Check":
            # 切换到Check界面时不需要特殊处理
            pass


if __name__ == "__main__":
    """
    运行函数
    """
    root = tk.Tk()
    # 计算屏幕中央的位置
    screenwidth = root.winfo_screenwidth()
    screenheight = root.winfo_screenheight()
    width = 600 * multiple
    heigh = 580 * multiple
    # 将窗口居中显示
    root.geometry('%dx%d+%d+%d' % (width, heigh, (screenwidth - width) / 2, (screenheight - heigh) / 2))
    app = MainApp(root)
    root.mainloop()
