#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：MP TOOL PROJECT 
@File    ：check.py
@Author  ：zxh
@Date    ：2025/3/19 16:21 
'''
import pandas as pd
import re
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from read_excel import *
from copy import copy


def check_sequence(df):
    """
    基于C列项目号检查顺序排列
    :param df: 处理过的二维表格
    :return: 错误值
    """
    errors = []
    cleaned_project_numbers = df.iloc[:, 2]
    glass = df.iloc[:, 7]

    only_name = cleaned_project_numbers.astype(str) + '_' + df.iloc[:, 3].astype(str) + '_' + df.iloc[:, 5].astype(
        str) + '_' + glass.astype(str)
    # pd.Series(only_name).to_csv('only_name.txt', index=False, header=False, encoding='utf-8')
    # print("only_name 已保存到 only_name.txt 文件")

    # 初始化序号
    expected_sequence = 1

    for idx, (project_number, sequence) in enumerate(zip(only_name, df.iloc[:, 0])):
        if pd.isna(sequence) or sequence == '' or sequence == 0:
            errors.append((idx + 2, '序号', "序号为空值"))
            if idx > 0 and project_number != only_name[idx - 1]:
                expected_sequence += 1
            continue

        # 如果当前项目号与前一个项目号不同，则序号应递增
        if idx > 0 and project_number != only_name[idx - 1]:
            expected_sequence += 1

        # 检查序号是否正确，确保将sequence转换为整数
        try:
            sequence_int = int(sequence)
        except ValueError:
            errors.append((idx + 2, '序号', f"序号 '{sequence}' 不是有效的整数"))
            continue

        # 检查序号是否正确
        if sequence_int != expected_sequence:
            errors.append((idx + 2, '序号', f"期望 {expected_sequence}, 实际 {sequence_int}"))

    return errors


def check_sequence_for_fix(df):
    """
    基于C列项目号检查顺序排列，并只返回相同项目号且相同错误的第一行错误，用于校正序号
    :param df: 处理过的二维表格
    :return: 错误值列表，每个错误值为 (行号, 列名, 错误信息)
    """
    errors = []
    cleaned_project_numbers = df.iloc[:, 2].astype(str) + '_' + df.iloc[:, 3].astype(str) + '_' + df.iloc[:, 5].astype(
        str) + '_' + df.iloc[:, 7].astype(str)
    expected_sequence = 1  # 初始化期望的序号

    # 用于存储上一个错误信息和项目号
    last_error = None
    last_project_number = None

    for idx, (project_number, sequence) in enumerate(zip(cleaned_project_numbers, df.iloc[:, 0])):
        # 如果当前项目号与前一个项目号不同，则期望序号递增
        if idx > 0 and project_number != cleaned_project_numbers[idx - 1]:
            expected_sequence += 1

        # 检查当前行的序号是否为空值或无效值
        if pd.isna(sequence) or sequence == '' or sequence == 0:
            # 如果当前项目号与上一个项目号相同，且错误信息相同，则跳过
            if project_number == last_project_number and current_error == last_error:
                continue
            current_error = (idx + 2, '序号', expected_sequence, "序号为空值")
            # 否则，记录当前错误
            errors.append(current_error)
            last_error = current_error
            last_project_number = project_number
            continue

        # 检查序号是否为有效整数
        try:
            sequence_int = int(sequence)
        except ValueError:
            # 如果当前项目号与上一个项目号相同，且错误信息相同，则跳过
            if project_number == last_project_number and current_error == last_error:
                continue
            current_error = (idx + 2, '序号', f"序号 '{sequence}' 不是有效的整数")
            # 否则，记录当前错误
            errors.append(current_error)
            last_error = current_error
            last_project_number = project_number
            continue

        # 检查序号是否正确
        if sequence_int != expected_sequence:
            # 如果当前项目号与上一个项目号相同，且错误信息相同，则跳过
            if project_number == last_project_number and current_error == last_error:
                continue
            current_error = (idx + 2, '序号', expected_sequence, sequence_int)
            # 否则，记录当前错误
            errors.append(current_error)
            last_error = current_error
            last_project_number = project_number

    return errors


def check_grade(column):
    """
    检查等级
    :param column:list中等级所在的那一列数据
    :return: 错误值
    """
    valid_chars = {'-', 'A', 'V', 'v', 'G', 'G1'}
    errors = []
    for idx, value in enumerate(column):
        if pd.isna(value):
            continue
        if value == 0:
            continue
        if value not in valid_chars:
            errors.append((idx + 2, '等级', value))
    return errors


def check_ic_model(column, valid_IC):
    """
    检查IC型号,对比compare文件列表
    :param column: list中IC所在的那一列数据
    :param valid_IC: 有效IC的对比
    :return: 错误值
    """
    errors = []
    for idx, value in enumerate(column):
        if pd.isna(value) or value == 0:
            errors.append((idx + 2, 'IC', "IC为空值"))
            continue
        if value == 7272 or value == 7302:
            continue
        if str(value) not in valid_IC:  # 将值转换为字符串进行比较
            errors.append((idx + 2, 'IC', value))
    return errors


def check_glass_model(column, valid_models):
    """
    检查玻璃型号，对比compare文件列表
    :param column: list中玻璃所在的那一列数据
    :param valid_models: 有效玻璃型号的对比
    :return: 错误值
    """
    errors = []
    for idx, value in enumerate(column):
        if pd.isna(value) or value == 0:
            errors.append((idx + 2, '玻璃', "玻璃为空值"))
            continue
        if value not in valid_models:
            errors.append((idx + 2, '玻璃', value))
    return errors


def check_interface(column):
    """
    检查接口定义,'SPI', 'I2C'两种
    :param column: list中接口所在的那一列数据
    :return: 错误值
    """
    valid_interfaces = {'SPI', 'I2C'}
    errors = []
    for idx, value in enumerate(column):
        if pd.isna(value) or value == 0:
            errors.append((idx + 2, '接口', "接口为空值"))
            continue
        if value not in valid_interfaces:
            errors.append((idx + 2, '接口', value))
    return errors


def check_flash(column, interface_column):
    """
    检查flash，对比接口即SPI->N,I2C->Y
    :param column: list中flash所在的那一列数据
    :param interface_column: list中接口所在的那一列数据
    :return:错误值
    """
    errors = []
    for idx, (interface, flash) in enumerate(zip(interface_column, column)):
        if pd.isna(interface) or pd.isna(flash) or interface == 0 or flash == 0:
            if interface == 0 and flash != 0:
                errors.append((idx + 2, 'flash', "接口为空值，无法判断"))
            if flash == 0:
                errors.append((idx + 2, 'flash', "flash为空值"))
            continue
        if not ((interface == 'SPI' and flash == 'N') or (interface == 'I2C' and flash == 'Y')):
            errors.append((idx + 2, 'flash', flash))
    return errors


def check_tool_version(df, valid_version):
    """
    检查TOOL版本,对比compare文件列表
    :param df:list中TOOL版本所在的那一列数据
    :param valid_version:有效版本对比
    :return:错误值
    """
    update_column = df.iloc[:, 15]
    site_column = df.iloc[:, 10]
    column = df.iloc[:, 13]
    errors = []
    for idx, (value, update_value, site_value) in enumerate(zip(column, update_column, site_column)):
        # 如果站点为 "整机"，则 TOOL 版本忽略
        if site_value == "整机":
            continue
        # 如果站点为模组，则判断
        elif site_value == "模组":
            # 如果更新说明有字符，则 TOOL 版本不能为空
            if pd.notna(update_value) and str(update_value).strip() != '' and update_value != 0:
                if pd.isna(value) or value == 0 or value == '':
                    errors.append((idx + 2, 'Tool版本号', "TOOL版本为空值（有更新说明）"))
                elif value not in valid_version:
                    errors.append((idx + 2, 'Tool版本号', value))
            # 如果更新说明为空，则 TOOL 版本可以为空
            else:
                if pd.notna(value) and value != 0 and value != '' and value not in valid_version:
                    errors.append((idx + 2, 'Tool版本号', value))
    return errors


def check_date_format(df):
    """
    检查发布日期格式,YYY.MM.DD或YYYY.M.D
    :param df:list中发布日期所在的那一列数据
    :return:错误值
    """
    pattern = re.compile(r'^\d{4}\.(0?[1-9]|1[0-2])\.(0?[1-9]|[12][0-9]|3[01])$')
    errors = []
    column = df.iloc[:, 16]
    update_column = df.iloc[:, 15]
    for idx, (value, update_value) in enumerate(zip(column, update_column)):
        # 如果更新说明有字符，则发布日期不能为空
        if pd.notna(update_value) and str(update_value).strip() != '' and update_value != 0:
            if pd.isna(value) or value == 0 or value == '':
                errors.append((idx + 2, '发布日期', "发布日期为空值（有更新说明）"))
            elif not pattern.match(str(value)):
                errors.append((idx + 2, '发布日期', value))
        # 如果更新说明为空，则发布日期可以为空
        else:
            if pd.notna(value) and value != 0 and value != '' and not pattern.match(str(value)):
                errors.append((idx + 2, '发布日期', value))
    return errors


def fix_sequence(file_path, errors):
    """
    用于序列修改
    :param file_path: 文件路径
    :param errors: 需要修改的数值信息，该数值取自check_sequence_for_fix(df)
    :return:
    """
    workbook = openpyxl.load_workbook(file_path)
    # sheet = workbook.active
    sheet_name = "MP Project List（internal）"
    if sheet_name not in workbook.sheetnames:
        print(f"工作表 '{sheet_name}' 不存在")
        return

    sheet = workbook[sheet_name]
    # 指定第一列
    target_column = 1
    # 获取目标单元格
    for idx in errors:
        target_row = idx[0]
        new_value = idx[2]

        target_cell = sheet.cell(row=target_row, column=target_column)

        # 检查目标单元格是否在合并单元格中
        is_merged = False
        for merged_range in sheet.merged_cells.ranges:
            # 检查目标单元格是否在当前合并范围内
            if (merged_range.min_row <= target_row <= merged_range.max_row and
                    merged_range.min_col <= target_column <= merged_range.max_col):
                is_merged = True
                # 找到合并区域的左上角单元格
                top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                # 修改左上角单元格的值
                top_left_cell.value = new_value
                break

        # 如果目标单元格不在合并单元格中，直接替换该单元格的值
        if not is_merged:
            target_cell.value = new_value

    # return workbook
    # # 保存文件
    output_file_path = 'modified_excel_file.xlsx'
    workbook.save(output_file_path)

    # print(f"文件已保存为 {output_file_path}")


# def check_same_project(df):
#     '''
#     检查是否有相同项目号
#     :param df: 数组
#     :return:
#     '''
#     errors = []
#     project_column = df.iloc[:, 2].str.upper()  #项目号
#     sequence_column = df.iloc[:, 0]  # 序号
#
#     project_dict = {}
#     project_prev_seq = {}
#
#     for idx, (project, seq) in enumerate(zip(project_column, sequence_column)):
#         row_num = idx + 2
#
#         # 跳过空值或无效项目号的行
#         if pd.isna(project) or project == '':
#             continue
#
#         # 跳过空序号的行
#         if pd.isna(seq) or seq == '':
#             continue
#
#         # 检查项目号是否已存在
#         if project in project_dict:
#             first_row, first_seq = project_dict[project]
#
#             # 只有当序号不同时才报错
#             if seq != first_seq and abs(seq - first_seq) != 1:
#                 errors.append((first_row, row_num, '项目号', project))
#
#         else:
#             # 记录项目号第一次出现的位置和序号
#             project_dict[project] = (row_num, seq)
#
#     return errors

# def check_same_project(df):
#     '''
#     检查是否有相同项目号
#     :param df: 数组
#     :return:
#     '''
#     errors = []
#     project_column = df.iloc[:, 2].str.upper()  # 项目号
#     sequence_column = df.iloc[:, 0]  # 序号
#
#     # 用于记录每个项目号的所有出现位置和序号
#     project_records = {}
#
#     for idx, (project, seq) in enumerate(zip(project_column, sequence_column)):
#         row_num = idx + 2
#
#         # 跳过空值或无效项目号的行
#         if pd.isna(project) or project == '':
#             continue
#
#         # 跳过空序号的行
#         if pd.isna(seq) or seq == '':
#             continue
#
#         # 检查项目号是否已存在
#         if project in project_records:
#             # 获取该项目号的所有历史记录
#             records = project_records[project]
#             invalid_records = []
#
#             # 检查当前序号是否与任何历史序号相同或相邻
#             is_valid = any(seq == prev_seq or abs(seq - prev_seq) == 1 for prev_seq, _ in records)
#
#             # 如果当前序号与任何历史序号都不相同且不相邻，则报错
#             if not is_valid:
#                 # 找到第一个不满足条件的记录作为错误参考
#                 for prev_seq, prev_row in records:
#                     if seq != prev_seq and abs(seq - prev_seq) != 1:
#                         errors.append((prev_row, row_num, '项目号', project))
#                         break
#         else:
#             # 初始化该项目号的记录列表
#             project_records[project] = []
#
#         # 将当前记录添加到项目号的历史记录中
#         project_records[project].append((seq, row_num))
#
#     return errors

def check_same_project(df):
    '''
    检查是否有相同项目号
    :param df: 数组
    :return: 错误列表，每个错误包含(行号1, 行号2, 字段名, 项目号)
    '''
    errors = []
    project_column = df.iloc[:, 2].str.upper()  # 项目号
    sequence_column = df.iloc[:, 0]  # 序号

    # 用于记录每个项目号的所有分组信息
    project_groups = {}

    for idx, (project, seq) in enumerate(zip(project_column, sequence_column)):
        row_num = idx + 2

        # 跳过空值或无效项目号的行
        if pd.isna(project) or project == '':
            continue

        # 跳过空序号的行
        if pd.isna(seq) or seq == '':
            continue

        # 检查项目号是否已存在
        if project in project_groups:
            # 获取该项目号的所有分组
            groups = project_groups[project]
            matched_group = None

            # 检查当前序号是否与任何分组中的序号相同或相邻
            for group in groups:
                if any(seq == prev_seq or abs(seq - prev_seq) == 1 for prev_seq, _ in group):
                    matched_group = group
                    break

            if matched_group is None:
                # 如果没有匹配的分组，创建新分组
                new_group = [(seq, row_num)]
                groups.append(new_group)
                # 为新分组的所有行与第一个分组的最后一行创建错误
                if groups[0]:  # 确保第一个分组不为空
                    last_row = groups[0][-1][1]  # 获取第一个分组的最后一行的行号
                    first_row = groups[0][0][1]
                    # 为当前行创建错误
                    errors.append((last_row, row_num, '项目号', project,first_row))
            else:
                # 如果匹配到分组，将当前行添加到该分组
                matched_group.append((seq, row_num))
                # 如果当前分组不是第一个分组，需要为当前行与第一个分组的最后一行创建错误
                if matched_group is not groups[0] and groups[0]:
                    last_row = groups[0][-1][1]  # 获取第一个分组的最后一行的行号
                    first_row = groups[0][0][1]
                    errors.append((last_row, row_num, '项目号', project,first_row))

        else:
            # 初始化该项目号的分组列表，第一个分组
            project_groups[project] = [[(seq, row_num)]]

    return errors


def cut_and_paste_excel(file_path, errors):
    """
    通过分割和重新组合的方式实现剪切粘贴功能
    :param file_path: Excel文件路径
    :param errors: 错误列表，格式为[(correct_row, error_row, field_name, project)]
    :return: 保存后的文件路径
    """
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    # ws = workbook.active
    sheet_name = "MP Project List（internal）"
    if sheet_name not in workbook.sheetnames:
        print(f"工作表 '{sheet_name}' 不存在")
        return
    ws = workbook[sheet_name]

    # 按项目号分组错误
    error_groups = {}
    for error in errors:
        correct_row, error_row, field_name, project, correct_start_row = error
        if project not in error_groups:
            error_groups[project] = {
                'correct_row': correct_row,
                'correct_start_row': correct_start_row,  # 正确区块的开始行
                'error_rows': []
            }
        error_groups[project]['error_rows'].append(error_row)

    # 收集所有需要处理的错误行范围和对应的正确行位置
    error_blocks = []
    for project, group_info in error_groups.items():
        error_rows = sorted(group_info['error_rows'])
        if error_rows:
            # 将连续的错误行合并为范围
            ranges = []
            start = error_rows[0]
            end = error_rows[0]

            for i in range(1, len(error_rows)):
                if error_rows[i] == error_rows[i - 1] + 1:
                    end = error_rows[i]
                else:
                    ranges.append((start, end))
                    start = error_rows[i]
                    end = error_rows[i]
            ranges.append((start, end))

            for start, end in ranges:
                error_blocks.append({
                    'range': (start, end),
                    'paste_after': group_info['correct_row'],
                    'correct_start_row': group_info['correct_start_row'],
                    'project': project
                })

    # 按起始行排序错误范围
    error_blocks.sort(key=lambda x: x['range'][0])

    print("检测到的错误区块:")
    for block in error_blocks:
        print(f"  项目 {block['project']}: {block['range'][0]}-{block['range'][1]} -> 插入到 {block['paste_after']} 之后")

    # 构建所有分割点
    total_rows = ws.max_row
    split_points = set()

    # 添加所有分割点：错误区块的起始和结束位置，以及正确行位置
    split_points.add(1)  # 开始
    split_points.add(total_rows + 1)  # 结束

    for block in error_blocks:
        start, end = block['range']
        split_points.add(start)
        split_points.add(end + 1)
        split_points.add(block['paste_after'] + 1)

    # 排序分割点
    split_points = sorted(split_points)

    # 根据分割点创建区块
    normal_blocks = []
    for i in range(len(split_points) - 1):
        start = split_points[i]
        end = split_points[i + 1] - 1
        if start <= end:
            normal_blocks.append({
                'range': (start, end),
                'type': 'normal'
            })

    # 创建最终的区块列表：正常区块 + 错误区块（在正确位置）
    all_blocks = []

    # 首先添加所有正常区块，但跳过包含错误行的区块
    for block in normal_blocks:
        start, end = block['range']
        is_error_block = False
        for error_block in error_blocks:
            error_start, error_end = error_block['range']
            if not (end < error_start or start > error_end):
                is_error_block = True
                break
        if not is_error_block:
            all_blocks.append(block)

    # 现在插入错误区块到正确位置
    for error_block in error_blocks:
        paste_after = error_block['paste_after']
        insert_index = 0
        for i, block in enumerate(all_blocks):
            start, end = block['range']
            if paste_after >= start and paste_after <= end:
                insert_index = i + 1
                break
            elif end < paste_after:
                insert_index = i + 1

        all_blocks.insert(insert_index, {
            'range': error_block['range'],
            'type': 'error',
            'project': error_block['project'],
            'paste_after': paste_after
        })

    print("\n最终区块顺序:")
    for i, block in enumerate(all_blocks, 1):
        start, end = block['range']
        if block['type'] == 'normal':
            print(f"{i}. 正常区块: {start}-{end}")
        else:
            print(f"{i}. 错误区块: {start}-{end} (项目: {block['project']})")

    # 创建新工作簿
    new_workbook = openpyxl.Workbook()
    new_ws = new_workbook.active
    new_ws.title = ws.title

    # 复制列宽
    for col in range(1, ws.max_column + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        if col_letter in ws.column_dimensions:
            new_ws.column_dimensions[col_letter].width = ws.column_dimensions[col_letter].width

    # 按新顺序复制数据
    current_new_row = 1

    # 记录每个区块在新文件中的位置
    block_new_positions = {}

    # 记录每个项目在新文件中的起始和结束行
    project_ranges = {}

    # 复制数据
    for block in all_blocks:
        start, end = block['range']
        rows_to_copy = end - start + 1
        new_start = current_new_row
        new_end = current_new_row + rows_to_copy - 1

        print(f"复制区块 {start}-{end} ({block['type']}) 到新文件的第 {new_start}-{new_end} 行")

        # 记录区块在新文件中的位置
        block_new_positions[(start, end)] = (new_start, new_end)

        # 复制数据和格式
        for row_idx in range(start, end + 1):
            for col in range(1, ws.max_column + 1):
                old_cell = ws.cell(row=row_idx, column=col)
                new_cell = new_ws.cell(row=current_new_row, column=col)
                new_cell.value = old_cell.value

                if old_cell.font:
                    new_cell.font = copy(old_cell.font)
                if old_cell.fill:
                    new_cell.fill = copy(old_cell.fill)
                if old_cell.border:
                    new_cell.border = copy(old_cell.border)
                if old_cell.alignment:
                    new_cell.alignment = copy(old_cell.alignment)
                new_cell.number_format = old_cell.number_format

            if row_idx in ws.row_dimensions:
                new_ws.row_dimensions[current_new_row].height = ws.row_dimensions[row_idx].height

            current_new_row += 1

    # 计算每个项目在新文件中的位置
    print("\n计算项目位置:")
    for project, group_info in error_groups.items():
        correct_start_row = group_info['correct_start_row']

        # 找到正确开始行在新文件中的位置
        new_correct_start = None
        for block_range, new_range in block_new_positions.items():
            old_start, old_end = block_range
            new_start, new_end = new_range
            if old_start <= correct_start_row <= old_end:
                offset = correct_start_row - old_start
                new_correct_start = new_start + offset
                break

        # 找到该项目错误区块在新文件中的结束位置
        new_error_end = None
        for error_block in error_blocks:
            if error_block['project'] == project:
                error_range = error_block['range']
                if error_range in block_new_positions:
                    new_error_start, new_error_end = block_new_positions[error_range]

        if new_correct_start and new_error_end:
            # 项目区间从正确开始行的新位置开始，到错误区块结束
            project_start = new_correct_start  # 这里不再+1
            project_end = new_error_end
            project_ranges[project] = {
                'start': project_start,
                'end': project_end
            }
            print(f"项目 {project}: 原正确开始行 {correct_start_row} -> 新位置 {new_correct_start}")
            print(f"  项目区间: {project_start}-{project_end}")

    # 处理合并单元格
    for merged_range in ws.merged_cells.ranges:
        try:
            start_cell = merged_range.min_row
            start_col = merged_range.min_col
            end_cell = merged_range.max_row
            end_col = merged_range.max_col

            def find_new_row(old_row):
                for block_range, new_range in block_new_positions.items():
                    old_start, old_end = block_range
                    new_start, new_end = new_range
                    if old_start <= old_row <= old_end:
                        offset = old_row - old_start
                        return new_start + offset
                return old_row

            new_start_row = find_new_row(start_cell)
            new_end_row = find_new_row(end_cell)

            start_col_letter = openpyxl.utils.get_column_letter(start_col)
            end_col_letter = openpyxl.utils.get_column_letter(end_col)

            new_merged_range = f"{start_col_letter}{new_start_row}:{end_col_letter}{new_end_row}"
            new_ws.merge_cells(new_merged_range)
            print(f"合并单元格: {start_cell}:{end_cell} -> {new_start_row}:{new_end_row}")

        except Exception as e:
            print(f"处理合并单元格时出错: {e}")
            continue

    # 生成合并范围列表
    merge_ranges = []
    for project, range_info in project_ranges.items():
        start = range_info['start']
        end = range_info['end']
        merge_range = f'C{start}:C{end}'
        merge_ranges.append(merge_range)
        print(f"项目 {project} 完整区间: {start}-{end} -> {merge_range}")

    # 保存文件
    try:
        save_path = "cut_paste_result.xlsx"
        new_workbook.save(save_path)
        print(f"文件已保存: {save_path}")
    except Exception as e:
        print(f"保存文件时出错: {e}")
        save_path = "temp_cut_paste_result.xlsx"
        new_workbook.save(save_path)
        print(f"文件已保存为: {save_path}")
    return save_path, merge_ranges



def set_merged_border(ws, min_row, min_col, max_row, max_col, original_border):
    """
    为合并后的单元格设置正确的边框,只保留外边框
    :param ws:
    :param min_row:
    :param min_col:
    :param max_row:
    :param max_col:
    :param original_border:
    :return:
    """
    # 定义边框样式 - 使用原边框或默认细边框
    left_border = original_border.left if original_border and original_border.left else Side(style='thin')
    right_border = original_border.right if original_border and original_border.right else Side(style='thin')
    top_border = original_border.top if original_border and original_border.top else Side(style='thin')
    bottom_border = original_border.bottom if original_border and original_border.bottom else Side(style='thin')

    # 设置合并区域所有单元格的边框（为取消可能存在的内部边框）
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)

            # 根据位置设置边框
            new_border = Border(
                left=left_border if col == min_col else None,  # 最左边列有左边框
                right=right_border if col == max_col else None,  # 最右边列有右边框
                top=top_border if row == min_row else None,  # 最上面行有上边框
                bottom=bottom_border if row == max_row else None  # 最下面行有下边框
            )

            cell.border = new_border


def merge_cells_like_excel(workbook_path, sheet_name, merge_ranges, output_path=None, center_aligned=True):
    """
    像Excel一样合并单元格，完全消除内部边框，可选居中对齐
    workbook_path: Excel文件路径
    sheet_name: 工作表名称
    merge_ranges: 要合并的单元格范围列表
    output_path: 输出文件路径
    center_aligned: 是否居中对齐（默认为True）
    """

    # 加载工作簿和工作表
    wb = openpyxl.load_workbook(workbook_path)

    if sheet_name not in wb.sheetnames:
        print(f"工作表 '{sheet_name}' 不存在")
        return

    ws = wb[sheet_name]

    # 处理每个合并范围
    for merge_range in merge_ranges:
        try:
            # 检查是否已经是合并单元格
            is_already_merged = False
            for merged_cell_range in ws.merged_cells.ranges:
                if str(merged_cell_range) == merge_range:
                    is_already_merged = True
                    print(f"范围 {merge_range} 已经是合并单元格，跳过")
                    break

            if not is_already_merged:
                # 获取范围边界
                min_col, min_row, max_col, max_row = range_boundaries(merge_range)

                # 保存左上角单元格的数据和格式
                top_left_cell = ws.cell(row=min_row, column=min_col)
                top_left_value = top_left_cell.value
                top_left_font = copy(top_left_cell.font)
                top_left_fill = copy(top_left_cell.fill)
                top_left_border = copy(top_left_cell.border)
                top_left_number_format = top_left_cell.number_format

                print(f"合并范围 {merge_range} - 左上角数据: {top_left_value}")

                # 先取消范围内所有单元格的合并状态（如果有）
                cells_to_unmerge = []
                for merged_cell_range in list(ws.merged_cells.ranges):
                    m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged_cell_range))
                    # 检查是否有重叠的合并单元格
                    if (m_min_row <= max_row and m_max_row >= min_row and
                            m_min_col <= max_col and m_max_col >= min_col):
                        cells_to_unmerge.append(str(merged_cell_range))

                for cell_range in cells_to_unmerge:
                    ws.unmerge_cells(cell_range)
                    print(f"  已取消合并: {cell_range}")

                # 执行合并
                ws.merge_cells(merge_range)

                # 获取合并后的单元格
                merged_cell = ws.cell(row=min_row, column=min_col)

                # 恢复数据和基本格式
                merged_cell.value = top_left_value
                merged_cell.font = top_left_font
                merged_cell.fill = top_left_fill
                merged_cell.number_format = top_left_number_format

                # 设置居中对齐（如果启用）
                if center_aligned:
                    merged_cell.alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
                else:
                    # 保持原来的对齐方式
                    if top_left_cell.alignment:
                        merged_cell.alignment = copy(top_left_cell.alignment)

                # 关键步骤：重新设置边框，只保留外边框，消除内部边框
                set_merged_border(ws, min_row, min_col, max_row, max_col, top_left_border)

                print(f"✓ 成功合并: {merge_range} (已消除内部边框)")

        except Exception as e:
            print(f"处理范围 {merge_range} 时出错: {str(e)}")

    # 保存文件
    if output_path is None:
        output_path = workbook_path.replace('.xlsx', '_excel_merged.xlsx')

    wb.save(output_path)
    print(f"文件已保存到: {output_path}")
    return output_path


def merge_and_center_cells(workbook_path, sheet_name, merge_ranges, output_path=None):
    """
    合并并居中单元格
    :param workbook_path:
    :param sheet_name:
    :param merge_ranges:
    :param output_path:
    :return:
    """
    return merge_cells_like_excel(
        workbook_path,
        sheet_name,
        merge_ranges,
        output_path,
        center_aligned=True
    )


# debug
# file_path='Project MP List&MP Flow List&CUT4兼容&HV noise (9).xlsx'
# sheet_name='MP Project List（internal）'
# df=read_excel_with_merged_cells(file_path, sheet_name)
#
# errors = check_same_project(df)
# cut_and_paste_excel(file_path, errors)

# for error in errors:
#     print(f"项目号 '{error[3]}' 在行 {error[0]} 和行 {error[1]} 重复")
